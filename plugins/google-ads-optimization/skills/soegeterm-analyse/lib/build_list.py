#!/usr/bin/env python3
"""Build the søgeterm-analyse deliverable: ONE flat colour-coded list (.xlsx).

Carl's call (2026-06-16): not the optimering-loop 8-tab workbook. ONE sheet, every term on one
blade, coloured by the model's verdict, with a short Danish reason — the way a Google Ads expert
actually reads a search-terms report. No overview/QS/skipped tabs. A small legend up top so the
colours are self-explanatory, and real Æ Ø Å throughout.

Input: the model's judged terms (slim rows + a `verdict` + `reason` the model added). Output: one
styled .xlsx. This builder makes NO judgement — it only renders what the model decided.

The five verdicts (the model assigns exactly one per term):
    VINDER       — converts / strong intent, not yet its own keyword -> promote
    RELEVANT     — on-offering, already well covered -> leave it
    NEGATIV      — should be blocked (off-offering, wrong intent, junk)
    FORKERT_PLACERET — relevant but in the wrong ad group / match -> restructure
    GRAENSE      — borderline, needs a human eye
"""
from __future__ import annotations

import re
import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

NAVY = "1F2A44"
NAVY_TEXT = "FFFFFF"
BAND = "F4F6FA"
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(bold=True, color=NAVY_TEXT, size=11)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BODY_FONT = Font(size=11, color="1F2A44")
_THIN = Side(style="thin", color="D6DBE6")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEAD_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP = Alignment(horizontal="left", vertical="top")

# One verdict -> one fill. Same pastel language as the rest of the suite so it reads consistently.
VERDICT_FILL = {
    "VINDER":           "D6EFD6",   # green
    "RELEVANT":         "D6E4F5",   # blue
    "FORKERT_PLACERET": "FCE4C8",   # orange
    "NEGATIV":          "F6D6D6",   # red
    "GRAENSE":          "E6E6E6",   # grey
}
VERDICT_LABEL = {
    "VINDER": "VINDER — konverterer / stærk intention, endnu ikke eget keyword → promovér",
    "RELEVANT": "RELEVANT — passer tilbuddet, allerede godt dækket → lad den være",
    "FORKERT_PLACERET": "FORKERT PLACERET — relevant, men forkert ad group / match → omstrukturér",
    "NEGATIV": "NEGATIV — bør blokeres (uden for tilbuddet, forkert intention, junk)",
    "GRAENSE": "GRÆNSE — grænsetilfælde, kræver et menneskeligt skøn",
}

# Main-sheet columns. Match type + Level sit on the main sheet so the auto-derived Negativ/Vinder
# sheets (and the future editor-csv-export bridge) have every Editor field they need without a
# second pass. Dom is the column the human edits; the two derived sheets FILTER on it.
COLUMNS = ["Søgeterm", "Foreslået keyword", "Foreslået match type", "Kampagne", "Ad group",
           "Triggerende keyword", "Keyword match type", "Budget brugt (DKK)", "Impressions",
           "Klik", "CTR (%)", "Konverteringer", "CPA (DKK)", "Level", "Allerede keyword?",
           "Dom", "Begrundelse"]
WRAP_COLS = {"Begrundelse"}
DOM_COL = "Dom"            # the column the derived sheets filter on
SUGGESTED_COL = "Foreslået keyword"   # the keyword that flows to Negativ/Vinder (need NOT == søgeterm)


def _fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")


def _already(v):
    if v is True:
        return "ja"
    if v is False:
        return "nej"
    return ""


def _filter_sheet(wb, title, src_title, data_first, data_last, verdict, col_map, constants=None):
    """Create a sheet whose rows are a live Google-Sheets FILTER of the main sheet, keeping only
    rows where the main sheet's Dom column == `verdict`. Updates automatically when the user edits
    Dom in Google Sheets (the chosen edit surface).

    title:     sheet name — use editor-csv-export's exact alias so the CSV bridge needs no rework.
    src_title: the main sheet's title (quoted in the formula).
    col_map:   [(editor_header, src_col_letter_or_None), ...] mapping main columns -> Editor names.
               A None src means the column is a constant (see constants).
    constants: {editor_header: literal} for non-column columns (e.g. Status="Paused").

    Each column gets ONE FILTER() in row 2 that spills down. The Dom condition column is looked up
    dynamically from COLUMNS (stays correct if the layout shifts). A constant column uses the SAME
    filter shape so it spills to identical height: FILTER(IF(<dom range>=v, "Add", ), <dom range>=v).
    """
    constants = constants or {}
    ws = wb.create_sheet(title)
    sq = src_title.replace("'", "''")           # escape single quotes for the sheet ref
    dom_col = get_column_letter(COLUMNS.index(DOM_COL) + 1)   # main-sheet Dom column letter
    dom_rng = f"'{sq}'!{dom_col}{data_first}:{dom_col}{data_last}"
    cond = f'{dom_rng}="{verdict}"'

    # A short note above the table so a human opening the sheet understands it is auto-derived.
    ws["A1"] = (f"Auto-genereret fra 'Søgetermer' (Dom = {verdict}). Ret Dom på hovedfanen, "
                f"så opdaterer denne sig selv i Google Sheets.")
    ws["A1"].font = Font(italic=True, size=10, color="606060")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(col_map), 2))

    header_row = 2
    for c, (hdr, _src) in enumerate(col_map, start=1):
        cell = ws.cell(row=header_row, column=c, value=hdr)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = HEAD_ALIGN; cell.border = BORDER
    ws.row_dimensions[header_row].height = 22

    formula_row = header_row + 1
    for c, (hdr, src) in enumerate(col_map, start=1):
        letter = get_column_letter(c)
        if src is None:                          # constant column, spilled to match height
            lit = constants.get(hdr, "")
            f = f'=FILTER(IF({cond},"{lit}",),{cond})'
        else:
            src_rng = f"'{sq}'!{src}{data_first}:{src}{data_last}"
            f = f'=FILTER({src_rng},{cond})'
        ws.cell(row=formula_row, column=c, value=f)

    widths = [26, 16, 22, 30, 14]
    for i in range(1, len(col_map) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i - 1 < len(widths) else 18
    ws.freeze_panes = ws.cell(row=formula_row, column=1).coordinate
    return ws


def _ngram_sheet(wb, ngram_rows, analysis=None):
    """Write the N-gram analysis tab: each n-gram's aggregated metrics across all terms containing
    it. Reference only (never an Editor import). Coloured by a SYSTEMIC heuristic so patterns pop:
    red = systemic waste (>=50 kr spend across the terms, 0 conversions); green = systemic winner
    (>=2 conversions); else neutral. The agent refines these calls — the colour is a starting read.
    The point: a word like 'gratis' that bleeds across 40 cheap terms shows up as ONE red row.

    analysis: optional written prose (the agent's overall read of the n-grams). Rendered as a styled
    box at the top — a navy 'Analyse'-bar + a bordered light panel, one wrapped row per paragraph
    (paragraphs split on blank lines). Sized to be pleasant to read for a longer write-up."""
    ws = wb.create_sheet("N-gram analyse")
    NCOLS = 11
    SYSTEMIC_WASTE = "F6D6D6"   # red:  spend across many terms, 0 conv
    SYSTEMIC_WIN = "D6EFD6"     # green: real conversions across the n-gram
    PANEL = "EEF1F7"            # soft panel background for the analysis box
    r = 1

    # --- styled Analyse-box (only if the agent supplied prose) ---
    if analysis and str(analysis).strip():
        bar = ws.cell(row=r, column=1, value="Analyse")
        bar.fill = HEADER_FILL; bar.font = Font(bold=True, size=12, color=NAVY_TEXT)
        bar.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        ws.row_dimensions[r].height = 24
        r += 1
        # one merged, wrapped row per paragraph; height scales with text length so nothing clips.
        paras = [p.strip() for p in re.split(r"\n\s*\n", str(analysis).strip()) if p.strip()]
        for p in paras:
            cell = ws.cell(row=r, column=1, value=p)
            cell.fill = _fill(PANEL); cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
            # ~ characters-per-line across the merged width (~140), ~16 px per wrapped line.
            lines = max(1, -(-len(p) // 140))
            ws.row_dimensions[r].height = max(18, lines * 16 + 6)
            r += 1
        ws.row_dimensions[r].height = 6  # spacer
        r += 1

    # --- intro note ---
    note = ws.cell(row=r, column=1, value=("N-gram analyse: hvert ord/frase aggregeret på tværs af "
            "ALLE søgetermer der indeholder det. Find systemisk spild og vindende temaer som "
            "enkelt-termer skjuler. Bloker/promovér ét n-gram i stedet for mange termer."))
    note.font = Font(italic=True, size=10, color="606060")
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    ws.row_dimensions[r].height = 30
    r += 1

    # --- colour legend (swatch in A, label spanning B:) ---
    ws.cell(row=r, column=1, value="Farvekoder:").font = BODY_FONT
    r += 1
    legend = [(SYSTEMIC_WASTE, "RØD — systemisk spild: ≥50 kr forbrug på tværs af termerne, 0 konverteringer → kandidat til at blokere n-grammet"),
              (SYSTEMIC_WIN, "GRØN — systemisk vinder OG udækket: konverterer, og termerne er endnu ikke keywords → ægte udvidelses-gab"),
              ("D6E4F5", "BLÅ — konverterer, men ≥80% af termerne er ALLEREDE keywords → dækket, ikke et udvidelses-gab (se 'Allerede dækket')"),
              (BAND, "NEUTRAL — hverken tydeligt spild eller vinder; vurdér i kontekst")]
    for hexv, label in legend:
        sw = ws.cell(row=r, column=1, value=""); sw.fill = _fill(hexv); sw.border = BORDER
        lab = ws.cell(row=r, column=2, value=label); lab.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
        r += 1
    r += 1  # spacer

    headers = ["N-gram", "Ord", "Antal termer", "Allerede dækket", "Budget brugt (DKK)",
               "Impressions", "Klik", "CTR (%)", "Konverteringer", "CPA (DKK)", "Konv.rate (%)",
               "Eksempel-termer"]
    hr = r   # table header sits below the analysis box + legend
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = HEAD_ALIGN; cell.border = BORDER
    ws.row_dimensions[hr].height = 24
    COVERED_WIN = "D6E4F5"   # blue: would be a winner, but already mostly covered -> not an expansion gap
    for i, g in enumerate(ngram_rows):
        rr = hr + 1 + i
        cost = _num_local(g.get("cost_dkk")); conv = _num_local(g.get("conversions"))
        covered_share = _num_local(g.get("covered_share_pct"))
        fill = None
        if conv == 0 and cost >= 50:
            fill = SYSTEMIC_WASTE                       # red: systemic waste
        elif conv >= 2:
            # a converting theme — but if it's already ~fully covered by keywords it's NOT an
            # expansion gap (Carl's point), so tone it blue instead of green.
            fill = COVERED_WIN if covered_share >= 80 else SYSTEMIC_WIN
        vals = [g.get("ngram", ""), g.get("words", ""), g.get("term_count", ""),
                g.get("covered_text", ""), g.get("cost_dkk", ""), g.get("impressions", ""),
                g.get("clicks", ""), g.get("ctr_pct", ""), g.get("conversions", ""),
                g.get("cpa_dkk", ""), g.get("conv_rate_pct", ""), g.get("example_terms", "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP if headers[c - 1] == "Eksempel-termer" else TOP
            if fill:
                cell.fill = _fill(fill)
            elif rr % 2 == 1:
                cell.fill = _fill(BAND)
    ws.freeze_panes = ws.cell(row=hr + 1, column=1).coordinate
    if ngram_rows:
        ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{hr + len(ngram_rows)}"
    for i, w in enumerate([26, 6, 12, 14, 16, 12, 8, 9, 13, 10, 12, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _num_local(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def build(data, out_path):
    """data = {client, account_id, period, today, scope, conversion_note, terms:[...], ngrams:[...]}
    Each term: slim fields + verdict + reason (+ optional suggested_keyword/match_type/level).
    ngrams (optional): ngram.analyse() output -> the N-gram analyse tab."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Søgetermer"

    client = data.get("client", "")
    terms = data.get("terms", [])

    # --- header band: title + context + legend, then the table. Row 1..N are meta, table follows.
    ws["A1"] = f"Søgeterm-analyse — {client}"
    ws["A1"].font = TITLE_FONT
    meta = [
        f"Konto: {data.get('account_id','')}    Periode: {data.get('period','')}    "
        f"Scope: {data.get('scope','hele kontoen')}    Genereret: {data.get('today','')}",
    ]
    if data.get("conversion_note"):
        meta.append(data["conversion_note"])
    meta.append("")
    meta.append("Farvekoder (dom pr. række):")
    r = 2
    for line in meta:
        ws.cell(row=r, column=1, value=line).font = BODY_FONT
        r += 1
    # legend swatches
    for key in ["VINDER", "RELEVANT", "FORKERT_PLACERET", "NEGATIV", "GRAENSE"]:
        sw = ws.cell(row=r, column=1, value="")
        sw.fill = _fill(VERDICT_FILL[key]); sw.border = BORDER
        lab = ws.cell(row=r, column=2, value=VERDICT_LABEL[key]); lab.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    r += 1  # spacer

    # --- table header ---
    header_row = r
    for c, h in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = HEAD_ALIGN; cell.border = BORDER
    ws.row_dimensions[header_row].height = 26

    # --- rows, whole-row coloured by verdict ---
    wrap_idx = {i for i, h in enumerate(COLUMNS, start=1) if h in WRAP_COLS}
    for i, t in enumerate(terms):
        rr = header_row + 1 + i
        verdict = str(t.get("verdict", "")).upper()
        fill = VERDICT_FILL.get(verdict)
        row_vals = {
            "Søgeterm": t.get("term", ""),
            # The keyword to actually add (negative or new). ONLY shown when the term is NOT already
            # a keyword (no point suggesting one that exists — Carl). Defaults to the search term,
            # but the agent/user may set something broader (e.g. 'helkropsscanning pris' ->
            # 'helkropsscanning'). The Negativ/Vinder sheets pull THIS, not the raw søgeterm.
            "Foreslået keyword": ("" if t.get("already_keyword")
                                  else (t.get("suggested_keyword") or t.get("term", ""))),
            "Kampagne": t.get("campaign", ""),
            "Ad group": t.get("ad_group", ""),
            "Triggerende keyword": t.get("trigger_keyword", ""),
            "Keyword match type": t.get("trigger_match_type", ""),
            "Budget brugt (DKK)": t.get("cost_dkk", ""),
            "Impressions": t.get("impressions", ""),
            "Klik": t.get("clicks", ""),
            "CTR (%)": t.get("ctr_pct", ""),
            "Konverteringer": t.get("conversions", ""),
            "CPA (DKK)": t.get("cpa_dkk", ""),
            # Editor-bound fields for the derived sheets + the future CSV bridge. Defaults are
            # sensible and editable: negatives default to Phrase + campaign level; a promoted
            # winner defaults to Exact. The agent may override per term in the judged JSON.
            # Match type for the SUGGESTED keyword (what gets added) — NOT the triggering keyword's
            # match type (that's the read-only "Keyword match type" column). This is what flows to
            # the Nye/Negative sheets -> Editor. Default Exact for a winner, Phrase for a negative.
            "Foreslået match type": t.get("match_type") or ("Exact" if verdict == "VINDER" else "Phrase"),
            "Level": t.get("level") or ("ad_group" if verdict == "VINDER" else "campaign"),
            "Allerede keyword?": _already(t.get("already_keyword")),
            "Dom": verdict,
            "Begrundelse": t.get("reason", ""),
        }
        for c, h in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=rr, column=c, value=row_vals[h])
            cell.border = BORDER
            cell.alignment = WRAP if c in wrap_idx else TOP
            if fill:
                cell.fill = _fill(fill)
            elif rr % 2 == 1:
                cell.fill = _fill(BAND)

    # --- widths + freeze the table header + autofilter over the table ---
    last = header_row + len(terms)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{max(last, header_row)}"
    widths = [28, 26, 20, 18, 22, 13, 13, 10, 7, 8, 12, 9, 11, 10, 13, 13, 44]
    for i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i - 1 < len(widths) else 16

    # --- the two auto-derived sheets (Sheets FILTER on the Dom column) ---
    # Named EXACTLY as editor-csv-export's tab aliases so the CSV bridge is zero-rework later, and
    # carrying the Editor column contract. Live FILTER() = they update when the user edits Dom in
    # Google Sheets. NOTE: openpyxl cannot READ a spilled FILTER result (formulas aren't evaluated
    # offline) — so the CSV bridge must RE-MATERIALISE these from the edited Dom column at export
    # time, reading the main sheet's hand-typed Dom cells (which openpyxl CAN read). That keeps the
    # sheet live for the human AND correct for the converter; it is not a silent trap.
    # Source column letters are derived from COLUMNS so they stay correct if the layout changes.
    def _col(name):
        return get_column_letter(COLUMNS.index(name) + 1)
    data_first = header_row + 1
    data_last = header_row + len(terms)

    # --- Negative keywords = Google Ads Editor's NEGATIVE-LIST bulk-upload template (Carl's screenshot 1).
    # Required: Action, Negative keyword list name OR Negative Keyword List ID, Negative keyword,
    # Keyword or list, Match type. The list NAME we cannot know -> a loud placeholder the user fills
    # (fails the import visibly rather than guessing). The keyword pulled is "Foreslået keyword"
    # (need not equal the søgeterm). Customer ID left blank (only needed for MCC multi-account upload).
    _filter_sheet(
        wb, "Negative keywords", ws.title, data_first, data_last, "NEGATIV",
        [("Action", None), ("Customer ID", None),
         ("Negative keyword list name", None), ("Negative Keyword List ID", None),
         ("Negative keyword", _col(SUGGESTED_COL)), ("Keyword or list", None),
         ("Match type", _col("Foreslået match type"))],
        constants={"Action": "Add", "Customer ID": "",
                   "Negative keyword list name": "<INDSÆT NEGATIVLISTE-NAVN>",
                   "Negative Keyword List ID": "", "Keyword or list": "keyword"},
    )
    # --- Nye keywords = Editor's ADD-KEYWORD template (screenshot 2). Required on create: Keyword;
    # required: Campaign + Ad group (by name). Action=Add, Keyword status=Paused (champion-challenger
    # safe). Keyword pulled is "Foreslået keyword".
    _filter_sheet(
        wb, "Nye keywords (vindere)", ws.title, data_first, data_last, "VINDER",
        [("Action", None), ("Keyword status", None),
         ("Campaign", _col("Kampagne")), ("Ad group", _col("Ad group")),
         ("Keyword", _col(SUGGESTED_COL)), ("Match Type", _col("Foreslået match type"))],
        constants={"Action": "Add", "Keyword status": "Paused"},
    )

    # --- N-gram analyse (systemiske mønstre på tværs af termer; reference, ikke import) ---
    if data.get("ngrams"):
        _ngram_sheet(wb, data["ngrams"], analysis=data.get("ngram_analysis"))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import argparse
    import json
    ap = argparse.ArgumentParser(description="Build the flat colour-coded søgeterm list (.xlsx).")
    ap.add_argument("--in", dest="inp", required=True, help="judged-terms JSON")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()
    build(json.loads(Path(args.inp).read_text()), args.out)
    print(f"Wrote {args.out}")
