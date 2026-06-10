#!/usr/bin/env python3
"""Build the optimization-loop REVIEW WORKBOOK (.xlsx) — the editable deliverable.

ARCHITECTURE (locked 2026-06-05): the loop does NOT emit Editor CSVs. It returns ONE
Excel workbook that the ads expert edits, can send to the client, and then hands to a
separate converter skill (in google-ads-general) that does workbook -> Editor CSV. This
mirrors the assembler, which was already made Excel-only (commit eb4ebd9): "the workbook
is a lossless superset; the converter produces the CSVs." Editor imports CSV, not .xlsx
(Google Ads Editor answer 30564) - which is exactly why the converter exists.

THE COLUMN CONTRACT (this is the interface the converter is built against):

Each entity tab splits its columns into two bands:
  1. EDITOR-BOUND columns - exact Google Ads Editor header spelling (answer 57747:
     headers are English, case/space-insensitive). The converter KEEPS these.
  2. METADATA columns - review context (reason, wasted spend, conversions, CPA, source).
     The converter DROPS these. They never go to Editor.

A sentinel row/format is not needed: the converter knows the fixed Editor header set per
entity (documented in SPEC section 3.5) and drops everything else.

### #Original — editing EXISTING entities (the loop's distinguishing need)

The assembler builds NEW campaigns (everything net-new, lands Paused, no #Original). The
loop optimises a LIVE account, so some rows EDIT an existing entity (e.g. an RSA rewrite).
Google Ads Editor uses the `<Column>#Original` convention (answer 57747) to match an edit
to the existing entity and preserve its history instead of creating a duplicate. So:
  - NET-NEW rows (new negative, promoted keyword, brand-new challenger RSA in an ad group
    that had none): no #Original columns. Status as appropriate.
  - EDIT rows (rewriting an existing RSA, changing an existing keyword): include the
    `#Original` column(s) carrying the current live value, so Editor edits in place.
The loop's execute stage decides per row which case applies; this builder writes whatever
#Original cells it is given and the converter preserves any `*#Original` column verbatim.

Self-bootstraps openpyxl. Runs locally and in an agent (the workflow JS itself cannot run
Python - see SPEC section 2).
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# --- Visual design system: SAME house look as the setup assembler workbook (assemble.py) ---
# Inbound dark-navy header, white bold text, soft zebra banding, thin grid, freeze + autofilter.
# The ONE optimizer-specific addition kept from before: a two-band header — Editor-bound columns
# get the full navy header (the converter KEEPS them), metadata columns get a lighter navy-tinted
# header (the converter DROPS them), so the expert can see at a glance which columns import.
NAVY = "1F2A44"          # Inbound dark navy (Editor-bound header)
NAVY_TEXT = "FFFFFF"
META_BG = "DCE3F0"       # lighter navy-tint header for metadata (converter-dropped) columns
META_TEXT = "1F2A44"
BAND_FILL = "F4F6FA"     # very light blue-grey zebra band for odd data rows
# Editor-bound DATA cells get a faint navy wash on EVERY row (not just the header) so the expert
# can see at a glance which columns import to Editor while scanning any row — Carl: "de først
# colloner ... er en anden farve så man ved hvad der indgår i CSV". Two shades keep the zebra
# rhythm under the wash (odd rows a touch deeper). Subtle enough to stay readable.
EDITOR_DATA_FILL = "E8EDF7"        # even data rows, Editor-bound columns
EDITOR_DATA_FILL_BAND = "DFE6F4"   # odd data rows, Editor-bound columns
TITLE_FONT = Font(bold=True, size=14, color=NAVY)

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
META_FILL = PatternFill(start_color=META_BG, end_color=META_BG, fill_type="solid")
BAND_PATTERN = PatternFill(start_color=BAND_FILL, end_color=BAND_FILL, fill_type="solid")
EDITOR_DATA_PATTERN = PatternFill(start_color=EDITOR_DATA_FILL, end_color=EDITOR_DATA_FILL, fill_type="solid")
EDITOR_DATA_BAND_PATTERN = PatternFill(start_color=EDITOR_DATA_FILL_BAND, end_color=EDITOR_DATA_FILL_BAND, fill_type="solid")
HEADER_FONT = Font(bold=True, color=NAVY_TEXT, size=11)
META_FONT = Font(bold=True, color=META_TEXT, size=11)
_THIN = Side(style="thin", color="D6DBE6")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_ALIGN = Alignment(horizontal="left", vertical="top")

# Metadata columns that hold prose and should wrap at a comfortable width (not autosize wide).
WRAP_COLS = {"Begrundelse", "Reason", "Niveau (oprindeligt)", "Konfidens-justering", "Flag"}
WIDE_WIDTH = 48
NARROW_MAX = 30

# --- Two colour axes (deliberately distinct; see references/selection-spec.md) ---
# BUCKET = classification (the "Alle søgetermer" overview tab colours by which group a term is in).
# BAND   = confidence/relevance (the Negative keywords tab colours by how safe a term is to block).
# Soft fills (pastel) so coloured rows stay readable under the house style.
BUCKET_FILLS = {
    "VINDER":            "D6EFD6",   # green   — winner, not yet a keyword
    "RELEVANT":          "D6E4F5",   # blue    — already covered / well-placed
    "PLACEMENT_PROBLEM": "FCE4C8",   # orange  — relevant, wrong ad group
    "IRRELEVANT":        "F6D6D6",   # red     — negative candidate
    "GRAENSE":           "E6E6E6",   # grey    — borderline
}
BAND_FILLS = {
    "GROEN": "D6EFD6",   # clearly off-offering -> safe to block
    "GUL":   "FBEFC8",   # loosely related -> check
    "ROED":  "F6D6D6",   # looks relevant -> probably should NOT be a negative
}
# Quality Score cell colour (1-10). Low = red (urgent), mid = orange, healthy = green. Keys are
# the string forms of the integer score (the colour helper compares str(value).upper()).
QS_FILLS = {
    "1": "F6D6D6", "2": "F6D6D6", "3": "FCE4C8", "4": "FCE4C8",
    "5": "FBEFC8", "6": "FBEFC8", "7": "EAF3E0", "8": "D6EFD6",
    "9": "D6EFD6", "10": "D6EFD6",
}
# QS component labels (creative / landing page / expected CTR). BELOW_AVERAGE = red so a cluster
# of below-average landing pages is visible at a glance; ABOVE = green; AVERAGE neutral.
QS_COMPONENT_FILLS = {
    "BELOW_AVERAGE": "F6D6D6",
    "AVERAGE":       "F4F6FA",
    "ABOVE_AVERAGE": "D6EFD6",
}

_MATCH = {"EXACT": "Exact", "PHRASE": "Phrase", "BROAD": "Broad"}

# --- Shared metric block (same header spelling across every tab; "ens" per Carl 2026-06-10) ---
# These metadata columns are IDENTICAL on Negative keywords, Nye keywords, and Alle søgetermer so
# the expert reads the same numbers the same way on every tab. The converter DROPS all of them
# (they are not Editor fields). Header spelling mirrors the standalone search-terms sheet Carl
# referenced. Negatives reframe the cost column as wasted spend (see METRIC_BLOCK_NEG) — that is
# the one deliberate per-tab difference.
METRIC_BLOCK = ["Budget brugt (DKK)", "Impressions", "Klik", "CTR (%)", "Konverteringer", "CPA (DKK)"]
METRIC_BLOCK_NEG = ["Spildt budget (DKK)", "Impressions", "Klik", "CTR (%)", "Konverteringer", "CPA (DKK)"]


def _match(raw):
    return _MATCH.get((raw or "").strip().upper(), raw or "")


def _fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")


def _ctr(clicks, impressions):
    """CTR (%) = clicks / impressions, guarded against ÷0. One decimal, blank if no impressions."""
    c, i = _coerce_num(clicks), _coerce_num(impressions)
    return round(c / i * 100, 1) if i else ""


def _cpa(cost_dkk, conversions):
    """CPA (DKK) = cost / conversions, guarded against ÷0. Whole DKK, blank if no conversions."""
    cost, conv = _coerce_num(cost_dkk), _coerce_num(conversions)
    return round(cost / conv, 0) if conv else ""


def _coerce_num(v):
    """Numbers arrive as int/float/str/None across the findings JSON; coerce safely to float."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _metric_block(row, cost_label="Budget brugt (DKK)"):
    """The shared metric dict for any tab. CTR and CPA are COMPUTED here (not stored upstream)
    so consistency is structural, not hand-maintained across tabs. cost_label lets the negatives
    tab show 'Spildt budget (DKK)' over the same underlying cost value."""
    cost = row.get("wasted_spend_dkk", row.get("cost_dkk", ""))
    clicks = row.get("clicks", "")
    impressions = row.get("impressions", "")
    conv = row.get("conversions", "")
    return {
        cost_label: cost,
        "Impressions": impressions,
        "Klik": clicks,
        "CTR (%)": _ctr(clicks, impressions),
        "Konverteringer": conv,
        "CPA (DKK)": row.get("cpa_dkk") if row.get("cpa_dkk") is not None else _cpa(cost, conv),
    }


def _sheet(wb, title, editor_headers, meta_headers, rows, widths=None):
    """Write one entity tab in the shared house style. editor_headers get the navy header
    (converter KEEPS); meta_headers get the lighter navy-tint header (converter DROPS). Zebra
    banding, thin grid, freeze header (A2), autofilter. Row 1 stays the header row (the converter
    reads headers from row 1 — never insert a banner row above the data)."""
    ws = wb.create_sheet(title)
    headers = editor_headers + meta_headers
    editor_set = set(editor_headers)
    # Header row (row 1).
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
        if h in editor_set:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = META_FILL
            cell.font = META_FONT
    ws.row_dimensions[1].height = 28
    # Body rows + zebra banding + borders + alignment. Editor-bound columns carry a faint navy
    # wash on every row (so the CSV columns stay visually distinct top to bottom); metadata
    # columns keep the plain zebra. The two Editor shades preserve the zebra rhythm.
    wrap_idx = {i for i, h in enumerate(headers, start=1) if h in WRAP_COLS}
    editor_idx = {i for i, h in enumerate(headers, start=1) if h in editor_set}
    for r, row in enumerate(rows, start=2):
        band = (r % 2 == 1)  # band odd data rows for subtle contrast
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.border = CELL_BORDER
            cell.alignment = WRAP_ALIGN if c in wrap_idx else TOP_ALIGN
            if c in editor_idx:
                cell.fill = EDITOR_DATA_BAND_PATTERN if band else EDITOR_DATA_PATTERN
            elif band:
                cell.fill = BAND_PATTERN
    # Freeze header + autofilter over the used range.
    ncols = len(headers)
    ws.freeze_panes = "A2"
    if ncols >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"
    # Column widths: explicit widths win; else wrap prose at WIDE_WIDTH, autosize the rest.
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[letter].width = widths[i - 1]
            continue
        if h in WRAP_COLS:
            ws.column_dimensions[letter].width = WIDE_WIDTH
            continue
        widest = len(str(h))
        for cell in ws[letter]:
            v = cell.value
            if v is None or str(v).startswith("="):
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = max(10, min(NARROW_MAX, widest + 3))
    return ws


def _reference_sheet(wb, title, headers, rows, row_fill_key=None, fill_map=None,
                     fill_column=None):
    """Write a REFERENCE tab (overview / sprunget-over) — never read by the converter, so it
    uses a single navy header (no editor/metadata split) and colours rows by a classification.

    title MUST NOT match any editor-csv-export tab alias (the converter reads keywords from
    'Keywords'/'Nye keywords (vindere)', negatives from 'Negative keywords'/'Negatives') — these
    reference tabs are named so they are structurally invisible to it (e.g. 'Alle søgetermer',
    'Sprunget over').

    row_fill_key: a key in each row dict whose value selects a colour from fill_map (e.g. the
        'bucket' or 'band'). When set, the WHOLE data row is filled with that colour.
    fill_column: if given, ONLY that column's cell is filled (used when colour belongs to one
        cell, e.g. a 'Konfidens' column), not the whole row.
    """
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.row_dimensions[1].height = 28
    wrap_idx = {i for i, h in enumerate(headers, start=1) if h in WRAP_COLS}
    for r, row in enumerate(rows, start=2):
        fill = None
        if row_fill_key and fill_map:
            fill = fill_map.get(str(row.get(row_fill_key, "")).upper())
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.border = CELL_BORDER
            cell.alignment = WRAP_ALIGN if c in wrap_idx else TOP_ALIGN
            if fill and (fill_column is None or h == fill_column):
                cell.fill = _fill(fill)
    ncols = len(headers)
    ws.freeze_panes = "A2"
    if ncols >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        if h in WRAP_COLS:
            ws.column_dimensions[letter].width = WIDE_WIDTH
            continue
        widest = len(str(h))
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = max(10, min(NARROW_MAX, widest + 3))
    return ws


def _color_column_by_value(ws, header_name, fill_map):
    """Fill each data cell in the named column by its own value (e.g. Konfidens -> band colour).
    Per-cell colouring, applied after _sheet so it overrides the zebra band on that one column."""
    headers = [c.value for c in ws[1]]
    if header_name not in headers:
        return
    col = headers.index(header_name) + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        fill = fill_map.get(str(cell.value or "").upper())
        if fill:
            cell.fill = _fill(fill)


# Editor-bound header sets per entity (the converter's KEEP list). Exact Editor spelling.
KEYWORDS_EDITOR = ["Campaign", "Ad group", "Keyword", "Match type", "Status"]
# Negatives speak the SAME vocabulary as the assembler's tab 04 so the converter's
# build_negatives() works unchanged for both workbooks (one contract, no fork):
# Campaign + Level + Ad group + Negative keyword + Match type. The converter derives
# Editor's Type (Campaign negative vs Negative) from Level, never from Match type.
NEGATIVES_EDITOR = ["Campaign", "Level", "Ad group", "Negative keyword", "Match type"]
RSA_EDITOR = (["Campaign", "Ad group", "Ad type", "Final URL", "Path 1", "Path 2"]
              + [f"Headline {i}" for i in range(1, 16)]
              + [f"Description {i}" for i in range(1, 5)]
              + ["Status"])


# --- Læs mig styling (document-like, not a data grid) ---
README_BODY_FONT = Font(size=11, color="1F2A44")
README_SECTION_FONT = Font(bold=True, size=12, color=NAVY_TEXT)
README_SECTION_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
README_SUBTLE_FILL = PatternFill(start_color="EEF1F7", end_color="EEF1F7", fill_type="solid")
README_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
README_LEFT_CENTER = Alignment(horizontal="left", vertical="center", indent=1)
README_LEGEND_LABEL = Alignment(horizontal="left", vertical="center", indent=1)


def _readme_sheet(wb, client, account_id, period, today, account_level_notes=None):
    """The 'Læs mig' tab as a styled one-page document (not a data grid): a navy title banner,
    navy section bars, breathing spacer rows, and a real colour legend with filled swatch cells
    next to their labels. Two colour axes are named DISTINCTLY (Carl keeps asking about colours):
      - BUCKET  = klassifikation, farver hele rækken på 'Alle søgetermer'.
      - KONFIDENS (bånd) = hvor trygt at blokere, farver Konfidens-cellen på 'Negative keywords'.
    Never read by the converter, so styled cells are safe. Real Æ Ø Å throughout (Carl: altid)."""
    ws = wb.create_sheet("Læs mig", 0)
    # Two columns: A = swatch / bullet gutter, B = text. Most text lives in B; section bars and
    # body paragraphs merge A:B so they read full-width like a document.
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 96

    r = 1

    def section(title):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=title)
        cell.fill = README_SECTION_FILL
        cell.font = README_SECTION_FONT
        cell.alignment = README_LEFT_CENTER
        ws.row_dimensions[r].height = 24
        r += 1

    def para(text, height=None, subtle=False):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = README_BODY_FONT
        cell.alignment = README_LEFT_TOP
        if subtle:
            cell.fill = README_SUBTLE_FILL
        if height:
            ws.row_dimensions[r].height = height
        r += 1

    def spacer(height=6):
        nonlocal r
        ws.row_dimensions[r].height = height
        r += 1

    def legend(hexv, label):
        """One legend line: a filled swatch in A, the label in B."""
        nonlocal r
        sw = ws.cell(row=r, column=1, value="")
        sw.fill = _fill(hexv)
        sw.border = CELL_BORDER
        lab = ws.cell(row=r, column=2, value=label)
        lab.font = README_BODY_FONT
        lab.alignment = README_LEGEND_LABEL
        r += 1

    # Title banner.
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    title = ws.cell(row=r, column=1, value=f"Optimerings-forslag — {client}")
    title.fill = HEADER_FILL
    title.font = Font(bold=True, size=15, color=NAVY_TEXT)
    title.alignment = README_LEFT_CENTER
    ws.row_dimensions[r].height = 34
    r += 1
    para(f"Konto: {account_id}        Periode: {period}        Genereret: {today}", subtle=True)
    spacer()

    # How to use.
    section("Sådan bruger du filen")
    para("1.  Gennemgå hver fane. De MØRKEBLÅ kolonner (yderst til venstre) er Google Ads "
         "Editor-felter — det er dem konverteringen sender videre. De LYSEBLÅ kolonner er "
         "kontekst til dig (begrundelse, spild, konverteringer, CTR) og bliver IKKE importeret.",
         height=46)
    para("2.  Ret frit: slet rækker du er uenig i, juster tekst, tilpas bud. Du bestemmer.",
         height=20)
    para("3.  Alle RSA-forslag er NYE challengers (status Paused) — ikke rettelser af "
         "eksisterende annoncer. Sæt den gamle annonce på pause eller fjern den FØRST når "
         "challengeren har vist sig bedre. (At redigere en live RSA nulstiller dens læring.)",
         height=46)
    para("4.  Når du er færdig: kør konverterings-skillet (workbook → Editor-CSV) og importér "
         "CSV'erne i Google Ads Editor (Account → Import → From file). Gennemgå grøn/gul diff, "
         "tryk Send.", height=32)
    spacer()

    # Colour legend — bucket axis.
    section("Farvekoder — fanen 'Alle søgetermer' (klassifikation)")
    para("Hele rækken farves efter hvilken GRUPPE søgetermet hører til:", height=18)
    legend(BUCKET_FILLS["VINDER"], "VINDER — konverterer (≥2) og er endnu ikke et keyword → foreslået som nyt keyword")
    legend(BUCKET_FILLS["RELEVANT"], "RELEVANT — passer tilbuddet og er allerede dækket / godt placeret")
    legend(BUCKET_FILLS["PLACEMENT_PROBLEM"], "PLACEMENT_PROBLEM — relevant, men i den forkerte ad group")
    legend(BUCKET_FILLS["IRRELEVANT"], "IRRELEVANT — passer ikke tilbuddet → negativ-kandidat")
    legend(BUCKET_FILLS["GRAENSE"], "GRÆNSE — grænsetilfælde, kræver et menneskeligt skøn")
    spacer()

    # Colour legend — confidence/band axis (named distinctly from bucket).
    section("Farvekoder — fanen 'Negative keywords' (konfidens / hvor trygt at blokere)")
    para("Her farves KONFIDENS-cellen — en ANDEN akse end grupperne ovenfor. Den siger hvor "
         "sikkert det er at blokere termet, forankret i hvad virksomheden faktisk tilbyder:",
         height=32)
    legend(BAND_FILLS["GROEN"], "GRØN — tydeligt uden for tilbuddet → trygt at blokere")
    legend(BAND_FILLS["GUL"], "GUL — løst relateret → tjek den manuelt før du blokerer")
    legend(BAND_FILLS["ROED"], "RØD — ser relevant ud → bør sandsynligvis IKKE være en negativ")
    para("Bemærk: 'Tynd data' er et SEPARAT flag (få klik), ikke en farve — et term kan være "
         "trygt at blokere OG have tyndt datagrundlag på samme tid.", height=32, subtle=True)
    spacer()

    # Guide to the gennemgang / diagnose tabs.
    section("Faner til gennemgang (bliver aldrig til CSV)")
    para("Disse faner er KUN til overblik og gennemgang — de eksporteres aldrig til Editor:",
         height=18)
    para("•  'Vindere til gennemgang' — søgetermer der KONVERTERER, men ser ud til at ligge uden "
         "for jeres tilbud. En konvertering er et lead, ikke bevis for at søgningen passede "
         "tilbuddet, så de bliver IKKE automatisk foreslået som nye keywords. Flyt en over til "
         "'Nye keywords' hvis du er enig i at den hører til.", height=46)
    para("•  'Quality Score' — de svageste keywords med deres tre QS-komponenter "
         "(annoncerelevans, landingsside, forventet CTR). Lave QS (1-2) og BELOW_AVERAGE-celler "
         "er røde, så en klynge med samme svaghed (fx samme landingsside) er let at se.",
         height=46)
    para("•  'Sprunget over' — ≥2-konv-termer der allerede er dækket af et eksisterende keyword "
         "(med angivelse af hvilket), så intet forsvinder usynligt.", height=32)
    para("•  'Alle søgetermer' — det fulde overblik over alle termer ≥5 DKK, farvet efter gruppe.",
         height=18)
    spacer()

    # Account-level negatives (only when present).
    if account_level_notes:
        section("Konto-niveau negative (vigtigt)")
        para("Editor kan ikke importere konto-niveau negative via CSV. Følgende ord er derfor "
             "udfoldet til en 'Campaign negative'-række PER aktiv kampagne på fanen 'Negative "
             "keywords'. Alternativt (renere): tilføj dem i stedet til en delt negativliste i "
             "Google Ads UI'en og tilknyt den til kontoens kampagner. Vælg det ene ELLER det "
             "andet:", height=60)
        for n in account_level_notes:
            para(f"     •  {n.get('keyword', '')}   ({n.get('reason', '')})", height=18)
        spacer()

    # Footer.
    para("Intet i denne fil er skrevet til kontoen. Du har fuld kontrol.", subtle=True)

    # Print setup: fit the wide text column to one page width so the document reads cleanly when
    # printed or exported to PDF (in Excel/Sheets column B scrolls fully regardless). Only the
    # width is forced — height is left to flow so long paragraphs are never clipped vertically.
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws


def build(data, out_path):
    """Build the review workbook.

    data schema:
    {
      "client", "account_id", "period", "today",
      "active_campaigns": ["string"],   # active campaign names; account-level negatives fan out across these
      "negatives": [ {keyword, match_type (EXACT|PHRASE), level (ad_group|campaign|account),
                      campaign, ad_group, wasted_spend_dkk, reason,
                      band (GROEN|GUL|ROED), clicks, impressions, thin_data (bool),
                      band_adjustment ("script: X -> agent: Y (grund: ...)")} ],
      "winners":   [ {term, campaign, ad_group, conversions, cpa_dkk, reason, agent_note} ],  # promote->Exact, Paused
      "review_winners": [ {term, campaign, ad_group, conversions, cost_dkk, clicks, impressions,
                            flag (off_offering), reason, agent_note} ],  # 'Vindere til gennemgang', never CSV
      "rsa_rows":  [ {campaign, ad_group, headlines[], descriptions[], paths[2], final_url,
                      status (Paused), reason} ],
      "all_terms": [ {term, campaign, ad_group, cost_dkk, clicks, conversions, cpa_dkk,
                      bucket (VINDER|RELEVANT|PLACEMENT_PROBLEM|IRRELEVANT|GRAENSE)} ],  # overview, never CSV
      "skipped_winners": [ {term, campaign, ad_group, conversions, cost_dkk,
                            covered_by:{keyword, match_type}} ],                        # 'Sprunget over', never CSV
      "quality_score": {                                                                # 'Quality Score', never CSV
        "average": float, "total_keywords": int,
        "worst": [ {campaign, ad_group, keyword, match_type, quality_score (1-10),
                    creative_quality, landing_page_quality, expected_ctr,
                    impressions, cost_dkk} ] }
    }
    Every RSA row is a NET-NEW challenger (Paused). The loop never edits a live RSA in place
    (resets learning + Editor CSV can't reliably match an RSA) — see the RSA tab comment.
    The `band`/`thin_data`/`bucket` fields come from lib/sweep.py's deterministic sweep; the
    agent may override `band` (logged in `band_adjustment`) and assigns `bucket`. Every metric
    field (cost_dkk, clicks, impressions, conversions, cpa_dkk) flows through from sweep's
    _row_metrics; the builder computes CTR (%) and missing CPA in _metric_block — see that helper.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    client = data.get("client", "")
    account_id = data.get("account_id", "")
    period = data.get("period", "")
    today = data.get("today", "")

    # --- Negative keywords tab ---
    # Google Ads Editor's CSV import supports ONLY campaign- and ad-group-level negatives
    # (Type = "Campaign negative" / "Negative", answer 57747). There is NO account-level
    # negative row in the CSV namespace; account-wide blocking is a shared-negative-list
    # (shared set) operation done in the UI. So an ACCOUNT-level finding is fanned out here
    # into one "Campaign negative" row per active campaign (same account-wide blocking effect,
    # fully importable) AND listed once on the Læs mig tab with the shared-list instruction.
    # The fan-out lives in the builder (not the converter) because the builder has the active
    # campaign list and because the workbook should be what-you-see-is-what-imports: the expert
    # reviews the actual per-campaign rows that will be created, not a single row that secretly
    # explodes at CSV time. Decision 2026-06-09.
    active_campaigns = data.get("active_campaigns") or []
    account_level_notes = []
    neg_rows = []

    def _neg_meta(n, zero_spend=False):
        """Confidence + significance + the shared metric block (converter DROPS all of these).
        Cost is framed as 'Spildt budget (DKK)' here — semantically right for a negative.
        zero_spend=True blanks the metric block (used on the 2nd+ fanned row of an account-level
        negative so the same spend is not counted once per campaign — see the fan-out below)."""
        block = _metric_block(n, cost_label="Spildt budget (DKK)")
        if zero_spend:
            block = {k: "" for k in block}
        return {
            "Konfidens": n.get("band", ""),
            "Tynd data": "ja" if n.get("thin_data") else "",
            "Konfidens-justering": n.get("band_adjustment", ""),
            **block,
            "Begrundelse": n.get("reason", ""),
        }

    for n in data.get("negatives", []):
        level = (n.get("level") or "campaign").lower()
        if level == "account":
            account_level_notes.append(n)
            targets = active_campaigns or [n.get("campaign", "")]
            # Account-level spend is one number; fanning it to N campaign rows would triple-count
            # it if every row repeated the cost. Show the real spend on the FIRST fanned row only;
            # the rest carry a blank metric block and a "samme spild" note. The blocking effect is
            # identical across all rows; only the reporting is de-duplicated.
            for j, camp in enumerate(targets):
                first = j == 0
                niveau = ("account -> fanned to campaign" if first
                          else "account -> fanned to campaign (samme spild som 1. række)")
                neg_rows.append({
                    "Campaign": camp,
                    "Level": "campaign",   # fanned to campaign-level rows
                    "Ad group": "",
                    "Negative keyword": n.get("keyword", ""),
                    "Match type": _match(n.get("match_type", "")),
                    "Niveau (oprindeligt)": niveau,
                    **_neg_meta(n, zero_spend=not first),
                })
            continue
        neg_rows.append({
            "Campaign": n.get("campaign", ""),
            "Level": level,
            "Ad group": n.get("ad_group", "") if level == "ad_group" else "",
            "Negative keyword": n.get("keyword", ""),
            "Match type": _match(n.get("match_type", "")),
            # metadata (dropped by converter):
            "Niveau (oprindeligt)": level,
            **_neg_meta(n),
        })
    # Readme first (inserted at index 0). Built here so it can list the account-level notes
    # computed in the negatives loop above.
    _readme_sheet(wb, client, account_id, period, today, account_level_notes)

    # No explicit widths: the shared style system autosizes short columns and wraps the prose
    # column (Begrundelse) at a comfortable width, exactly like the setup assembler workbook.
    # Metadata band carries the confidence/significance columns (all dropped by the converter).
    neg_meta_headers = (["Niveau (oprindeligt)", "Konfidens", "Tynd data", "Konfidens-justering"]
                        + METRIC_BLOCK_NEG + ["Begrundelse"])
    neg_ws = _sheet(wb, "Negative keywords", NEGATIVES_EDITOR, neg_meta_headers, neg_rows)
    # Colour the Konfidens CELL by band (the relevance axis: GROEN/GUL/ROED). Per-cell, not whole
    # row, so it doesn't fight the zebra banding on the other columns.
    _color_column_by_value(neg_ws, "Konfidens", BAND_FILLS)

    # --- Keyword expansion tab (promote winners to Exact, Paused) ---
    # Same shared metric block as every other tab (Carl: metrics "ens" across faner). The winner
    # rows carry the full metric set from sweep_winners (via _row_metrics), so the block is fully
    # populated — unlike the negatives tab, Konverteringer here is >=2 by construction.
    kw_rows = []
    for w in data.get("winners", []):
        kw_rows.append({
            "Campaign": w.get("campaign", ""),
            "Ad group": w.get("ad_group", ""),
            "Keyword": w.get("term", ""),
            "Match type": "Exact",
            "Status": "Paused",
            # metadata (dropped by converter):
            **_metric_block(w),
            "Begrundelse": w.get("reason", ""),
            "Agent-note": w.get("agent_note", ""),
        })
    _sheet(wb, "Nye keywords (vindere)", KEYWORDS_EDITOR,
           METRIC_BLOCK + ["Begrundelse", "Agent-note"], kw_rows)

    # --- Vindere til gennemgang — converting terms that look OFF-OFFERING (reference only) ---
    # The offering-grounded winner sweep (sweep.sweep_winners) splits >=2-conv novel terms by the
    # offering check: on-offering -> Nye keywords (promotable), off-offering -> HERE. A conversion
    # on a lead-gen account is a LEAD, not proof the search intent matched the offering (someone can
    # land and sign up for something else), so an off-offering destination like 'zanzibar højskole'
    # is SURFACED + flagged for the expert, never auto-promoted into a keyword. Named so it matches
    # NO editor-csv-export alias -> never becomes a CSV. The expert moves a confirmed one to the Nye
    # keywords tab by hand. See references/selection-spec.md (decision: flag, don't gate).
    review_winners = data.get("review_winners", [])
    if review_winners:
        rw_rows = [{
            "Søgeterm": w.get("term", ""),
            "Kampagne": w.get("campaign", ""),
            "Ad group": w.get("ad_group", ""),
            **_metric_block(w),
            "Flag": "Ser ud til at være uden for tilbuddet (off-offering)",
            "Begrundelse": w.get("reason", ""),
            "Agent-note": w.get("agent_note", ""),
        } for w in review_winners]
        _reference_sheet(wb, "Vindere til gennemgang",
                         ["Søgeterm", "Kampagne", "Ad group"] + METRIC_BLOCK
                         + ["Flag", "Begrundelse", "Agent-note"], rw_rows)

    # --- RSA challengers tab ---
    # EVERY RSA change is a NET-NEW challenger (Paused), never an in-place edit. Rationale
    # (decision 2026-06-09): editing a live RSA's creative resets its learning — RSAs are
    # effectively immutable in Google Ads, so "improving" one is a new-ad operation under the
    # hood (SPEC §6.4 + the ASA/Inbound memory note). Google Ads Editor's docs do not even
    # confirm RSA text-edit-in-place via CSV, and never enumerate which #Original fields would
    # match an RSA — so an #Original RSA edit row risks BOTH a silent duplicate (no match) and
    # clobbered headlines 2-15 (treated as full new content). The safe, idiomatic move is a
    # fresh challenger: the human pauses/removes the old ad after the challenger proves out.
    # (The converter still supports #Original passthrough for genuinely-editable entities like
    # keyword bid/URL — it is just never emitted for RSAs.)
    rsa_editor_headers = list(RSA_EDITOR)
    rsa_rows = []
    for r in data.get("rsa_rows", []):
        row = {
            "Campaign": r.get("campaign", ""),
            "Ad group": r.get("ad_group", ""),
            "Ad type": "Responsive search ad",
            "Final URL": r.get("final_url", ""),
            "Path 1": (r.get("paths") or ["", ""])[0],
            "Path 2": (r.get("paths") or ["", ""])[1] if len(r.get("paths") or []) > 1 else "",
            "Status": r.get("status", "Paused"),
        }
        for i, h in enumerate(r.get("headlines", [])[:15], start=1):
            row[f"Headline {i}"] = h
        for i, d in enumerate(r.get("descriptions", [])[:4], start=1):
            row[f"Description {i}"] = d
        row["Begrundelse"] = r.get("reason", "")
        rsa_rows.append(row)
    _sheet(wb, "RSA challengers", rsa_editor_headers, ["Begrundelse"], rsa_rows)

    # --- Sprunget over (vindere) — >=2-conv terms a script reason filtered out (reference only) ---
    # Named so it matches NO editor-csv-export alias -> never becomes a CSV. Answers "why did this
    # >=2-conv term not appear as a new keyword?" with the exact covering keyword.
    skipped = data.get("skipped_winners", [])
    if skipped:
        sk_rows = [{
            "Søgeterm": s.get("term", ""),
            "Kampagne": s.get("campaign", ""),
            "Ad group": s.get("ad_group", ""),
            "Konverteringer": s.get("conversions", ""),
            "Cost (DKK)": s.get("cost_dkk", ""),
            "Sprunget over fordi": "Allerede dækket af eksisterende keyword",
            "Dækket af (keyword)": (s.get("covered_by") or {}).get("keyword", ""),
            "Match type": (s.get("covered_by") or {}).get("match_type", ""),
        } for s in skipped]
        _reference_sheet(wb, "Sprunget over",
                         ["Søgeterm", "Kampagne", "Ad group", "Konverteringer", "Cost (DKK)",
                          "Sprunget over fordi", "Dækket af (keyword)", "Match type"], sk_rows)

    # --- Alle søgetermer — full overview of every term >=5 DKK, coloured by BUCKET (reference) ---
    # Named so it matches NO editor-csv-export alias -> never becomes a CSV. The action tabs above
    # are distilled subsets of this. Colour axis here is the classification bucket (NOT confidence).
    all_terms = data.get("all_terms", [])
    if all_terms:
        # Same shared metric block as the action tabs (Carl: "ens" across negative/nye/alle), then
        # the bucket group at the end. Whole-row coloured by Gruppe (the classification axis).
        ov_rows = [{
            "Søgeterm": t.get("term", ""),
            "Kampagne": t.get("campaign", ""),
            "Ad group": t.get("ad_group", ""),
            **_metric_block(t),
            "Gruppe": t.get("bucket", ""),
        } for t in all_terms]
        _reference_sheet(wb, "Alle søgetermer",
                         ["Søgeterm", "Kampagne", "Ad group"] + METRIC_BLOCK + ["Gruppe"], ov_rows,
                         row_fill_key="Gruppe", fill_map=BUCKET_FILLS)

    # --- Quality Score — flagged KEYWORDS (reference only, never a CSV) ---
    # QS is a first-class diagnosis but NOT a Google Ads Editor entity, so it gets a reference tab
    # named to match NO editor-csv-export alias. Grain is KEYWORD (there is no native ad-group QS —
    # see lib/gaql/quality_score.py). The QS cell is coloured by score so the actionable cluster
    # (QS 1-2 keywords, usually with a BELOW_AVERAGE landing page) is visible at a glance. Landing
    # page is shown as the API's component LABEL, never converted into a fabricated score.
    qs = data.get("quality_score") or {}
    qs_worst = qs.get("worst") or []
    if qs_worst:
        qs_rows = [{
            "Søgeterm (keyword)": k.get("keyword", ""),
            "Kampagne": k.get("campaign", ""),
            "Ad group": k.get("ad_group", ""),
            "Match type": _match(k.get("match_type", "")),
            "Quality Score": k.get("quality_score", ""),
            "Annoncerelevans": k.get("creative_quality", ""),
            "Landingsside": k.get("landing_page_quality", ""),
            "Forventet CTR": k.get("expected_ctr", ""),
            "Impressions": k.get("impressions", ""),
            "Budget brugt (DKK)": k.get("cost_dkk", k.get("cost", "")),
        } for k in qs_worst]
        qs_ws = _reference_sheet(wb, "Quality Score",
                                 ["Søgeterm (keyword)", "Kampagne", "Ad group", "Match type",
                                  "Quality Score", "Annoncerelevans", "Landingsside",
                                  "Forventet CTR", "Impressions", "Budget brugt (DKK)"], qs_rows)
        _color_column_by_value(qs_ws, "Quality Score", QS_FILLS)
        # Component-label columns coloured by BELOW/AVERAGE/ABOVE so a red landing-page cluster pops.
        for comp in ("Annoncerelevans", "Landingsside", "Forventet CTR"):
            _color_column_by_value(qs_ws, comp, QS_COMPONENT_FILLS)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import argparse
    import json
    ap = argparse.ArgumentParser(description="Build the optimization-loop review workbook (.xlsx).")
    ap.add_argument("--in", dest="inp", required=True, help="path to the findings JSON")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()
    build(json.loads(Path(args.inp).read_text()), args.out)
    print(f"Wrote {args.out}")
