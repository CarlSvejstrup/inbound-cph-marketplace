#!/usr/bin/env python3
"""OPT-IN: build a static, colour-coded Excel overview of the search-term review.

WHEN THIS RUNS
    Only when the expert asks for an Excel overview at the END (Trin 7C) — never in the normal flow.
    It is a human-readable REPORT/record artifact (send to a client, keep on file), distinct from the
    Editor-CSV (an import file) and the live MCP apply (the action). Same data, three different needs.

WHAT IT MAKES (one .xlsx, three tabs, all STATIC values — no live FILTER formulas, deliberately;
the live-formula version was fragile to maintain):
    1. "Søgetermer"  — the full judged term list, colour-coded by what the conversation DECIDED:
                       red = became a negative, green = promoted to a new keyword, grey = reviewed/left.
    2. "Beslutninger"— the agreed negatives and new keywords, fully concrete.
    3. "Temaer"      — the n-gram waste/winner themes + intent lenses from the digest (the interesting
                       systemic patterns), so the workbook also carries the "why".

NOT A JUDGE. The colour on a term is just a RECONCILIATION of the already-made decisions back onto the
list (does this term match a keyword we decided to block/add?) — display only, never a new verdict.

INPUTS
    --raw        the saved search-terms pull (report/GAQL rows, {results:[...]}, or slim output)
    --decisions  the agreed decisions.json (same file Trin 7A/7B use)
    --out        output .xlsx path

Reuses slim.py + ngram.py so the list + themes match the rest of the skill exactly.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import slim as slim_mod  # noqa: E402
import ngram as ngram_mod  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ABORT: openpyxl mangler. Installér med:  pip install openpyxl --break-system-packages")

# --- palette (tasteful, readable) --------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEG_FILL = PatternFill("solid", fgColor="F4CCCC")      # light red  — became a negative
WIN_FILL = PatternFill("solid", fgColor="D9EAD3")      # light green — promoted to keyword
WASTE_FILL = PatternFill("solid", fgColor="FCE5CD")    # light orange — waste theme
NEUTRAL_FONT = Font(color="666666")
BOLD = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=False)


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _tokens(s: str) -> list:
    return [t for t in re.split(r"[^0-9a-zæøå]+", str(s or "").lower()) if t]


def _matches(term: str, keyword: str, match_type: str) -> bool:
    """Does `term` fall under the decided `keyword` at `match_type`? Display reconciliation only —
    mirrors how Google would match, roughly: exact = same text; phrase = contiguous token run;
    broad = all keyword tokens present somewhere in the term."""
    t, kw = str(term or "").lower().strip(), str(keyword or "").lower().strip()
    if not t or not kw:
        return False
    mt = (match_type or "").lower()
    if mt == "exact":
        return t == kw
    if mt == "phrase":
        return re.search(r"(^|\s)" + re.escape(kw) + r"(\s|$)", t) is not None
    # broad (and anything else): all keyword tokens appear in the term
    tt = set(_tokens(t))
    return all(w in tt for w in _tokens(kw))


def _status(term: str, negatives: list, new_keywords: list) -> str:
    """Reflect the conversation's decisions onto a term. Negative wins if both somehow match."""
    for n in negatives:
        if _matches(term, n.get("keyword"), n.get("match_type")):
            return "negativ"
    for k in new_keywords:
        if _matches(term, k.get("keyword"), k.get("match_type")):
            return "tilføjet"
    return "gennemgået"


def _autosize(ws, max_w=60):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(10, width + 2))


def _header(ws, row_idx, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP


def _tab_terms(wb, terms, negatives, new_keywords):
    ws = wb.active
    ws.title = "Søgetermer"
    headers = ["Term", "Forbrug (kr)", "Klik", "Impr.", "Konv.", "CPA (kr)", "CTR %", "Ad group",
               "Kampagne", "Triggerende kw", "Match type", "Allerede kw?", "Status"]
    _header(ws, 1, headers)
    n_neg = n_win = 0
    for t in sorted(terms, key=lambda r: -_num(r.get("cost_dkk"))):
        st = _status(t.get("term", ""), negatives, new_keywords)
        already = t.get("already_keyword")
        already_txt = "" if already is None else ("Ja" if already else "Nej")
        st_txt = {"negativ": "→ Negativ", "tilføjet": "→ Tilføjet keyword", "gennemgået": "Gennemgået"}[st]
        rowvals = [t.get("term", ""), round(_num(t.get("cost_dkk")), 0), int(_num(t.get("clicks"))),
                   int(_num(t.get("impressions"))), round(_num(t.get("conversions")), 1),
                   t.get("cpa_dkk", ""), t.get("ctr_pct", ""), t.get("ad_group", ""),
                   t.get("campaign", ""), t.get("trigger_keyword", ""),
                   (t.get("trigger_match_type") or t.get("keyword_match_type") or ""),
                   already_txt, st_txt]
        r = ws.max_row + 1
        for j, v in enumerate(rowvals, start=1):
            ws.cell(row=r, column=j, value=v)
        fill = {"negativ": NEG_FILL, "tilføjet": WIN_FILL}.get(st)
        if fill:
            ws.cell(row=r, column=13).fill = fill
            if st == "negativ":
                n_neg += 1
            else:
                n_win += 1
        else:
            ws.cell(row=r, column=13).font = NEUTRAL_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _autosize(ws)
    return n_neg, n_win


def _tab_decisions(wb, negatives, new_keywords):
    ws = wb.create_sheet("Beslutninger")
    ws.cell(row=1, column=1, value="NEGATIVES").font = BOLD
    _header(ws, 2, ["Keyword", "Match type", "Level", "Kampagne", "Ad group / liste"])
    for n in negatives:
        target = n.get("ad_group") or n.get("list_name") or ""
        ws.append(["", "", "", "", ""])  # placeholder to advance; overwrite below
        r = ws.max_row
        for j, v in enumerate([n.get("keyword", ""), (n.get("match_type") or "").capitalize(),
                               (n.get("level") or "campaign"), n.get("campaign", ""), target], start=1):
            ws.cell(row=r, column=j, value=v)
            ws.cell(row=r, column=j).fill = NEG_FILL
    gap = ws.max_row + 2
    ws.cell(row=gap, column=1, value="NYE KEYWORDS").font = BOLD
    _header(ws, gap + 1, ["Keyword", "Match type", "Kampagne", "Ad group", "Status v. live"])
    for k in new_keywords:
        ws.append(["", "", "", "", ""])
        r = ws.max_row
        for j, v in enumerate([k.get("keyword", ""), (k.get("match_type") or "").capitalize(),
                               k.get("campaign", ""), k.get("ad_group", ""), "Aktiv (ingen paused via MCP)"],
                              start=1):
            ws.cell(row=r, column=j, value=v)
            ws.cell(row=r, column=j).fill = WIN_FILL
    _autosize(ws)


def _tab_themes(wb, terms):
    ws = wb.create_sheet("Temaer")
    ngrams = ngram_mod.analyse(terms)
    waste = sorted([g for g in ngrams if g.get("term_count", 0) >= 2 and _num(g.get("cost_dkk")) >= 50
                    and _num(g.get("conversions")) == 0], key=lambda g: -_num(g.get("cost_dkk")))
    winners = sorted([g for g in ngrams if g.get("term_count", 0) >= 2 and _num(g.get("conversions")) > 0],
                     key=lambda g: -_num(g.get("conversions")))
    ws.cell(row=1, column=1, value="SYSTEMISKE SPILD-TEMAER (mange termer · 0 konv) — samtalestof").font = BOLD
    _header(ws, 2, ["N-gram", "Termer", "Forbrug (kr)", "Konv.", "Eksempler"])
    for g in waste[:20]:
        ws.append([g.get("ngram"), g.get("term_count"), round(_num(g.get("cost_dkk")), 0),
                   round(_num(g.get("conversions")), 1), g.get("example_terms", "")])
        for j in range(1, 6):
            ws.cell(row=ws.max_row, column=j).fill = WASTE_FILL
    gap = ws.max_row + 2
    ws.cell(row=gap, column=1, value="VINDER-TEMAER (konverterer på tværs af termer)").font = BOLD
    _header(ws, gap + 1, ["N-gram", "Termer", "Forbrug (kr)", "Konv.", "Dækket", "Eksempler"])
    for g in winners[:20]:
        ws.append([g.get("ngram"), g.get("term_count"), round(_num(g.get("cost_dkk")), 0),
                   round(_num(g.get("conversions")), 1), g.get("covered_text", ""), g.get("example_terms", "")])
        for j in range(1, 7):
            ws.cell(row=ws.max_row, column=j).fill = WIN_FILL
    _autosize(ws)


def build(terms, negatives, new_keywords, out_path: str) -> dict:
    wb = Workbook()
    n_neg, n_win = _tab_terms(wb, terms, negatives, new_keywords)
    _tab_decisions(wb, negatives, new_keywords)
    _tab_themes(wb, terms)
    wb.save(out_path)
    return {"file": out_path, "terms": len(terms), "negatives": len(negatives),
            "new_keywords": len(new_keywords), "terms_marked_negative": n_neg,
            "terms_marked_added": n_win}


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Build the opt-in colour-coded search-term Excel overview.")
    ap.add_argument("--raw", required=True, help="saved pull: report/GAQL rows, {results:[...]}, or slim output")
    ap.add_argument("--decisions", required=True, help="agreed decisions.json")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    payload = json.loads(Path(args.raw).read_text())
    if isinstance(payload, dict) and "terms" in payload:
        terms = payload["terms"]
    else:
        raw = payload.get("results", payload) if isinstance(payload, dict) else payload
        terms = slim_mod.slim(raw, aggregate=True)["terms"]

    dec = json.loads(Path(args.decisions).read_text())
    res = build(terms, dec.get("negatives") or [], dec.get("new_keywords") or [], args.out)
    print(json.dumps(res, ensure_ascii=False, indent=2))
