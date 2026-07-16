#!/usr/bin/env python3
"""OPT-IN: build a static, colour-coded Excel overview of the Shopping-performance review.

WHEN THIS RUNS
    Only when the expert asks for an Excel overview (Trin 6 → spor B) — never in the normal flow.
    It is a human-readable REPORT/record artifact (send to a client, keep on file), distinct from
    the live MCP apply (the action). Same data, two different needs.

WHAT IT MAKES (one .xlsx, three tabs, all STATIC values):
    1. "Produkter"   — the full product list (aggregated across campaigns), colour-coded:
                       red = spend with 0 conv AND not PMax (talk about it),
                       green = converting product, grey = PMax product (product-grain conv unreliable).
    2. "Spild"       — Shopping products with spend and 0 conv, ranked by spend (PMax excluded).
    3. "ImpressionShare" — per-campaign IS with budget-lost vs rank-lost split.

NOT A JUDGE. The colour is a RECONCILIATION of what the digest computed back onto the list — display
only, never a new verdict. The cardinal rule holds: 0 conv is not proof of waste.

INPUTS
    --products   the saved pull A json (product performance)
    --is         the saved pull B json (campaign impression share)
    --out        output .xlsx path

Reuses slim.py + digest.py so the tables match the brief exactly.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import slim as slim_mod  # noqa: E402
import digest as digest_mod  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ABORT: openpyxl mangler. Installér med:  pip install openpyxl --break-system-packages")

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")  # Inbound-ish dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED = PatternFill("solid", fgColor="F8D7DA")
GREEN = PatternFill("solid", fgColor="D4EDDA")
GREY = PatternFill("solid", fgColor="E9ECEF")


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header(ws, cols: list[str]) -> None:
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def build(products: list[dict], campaigns: list[dict], out: str) -> None:
    d = digest_mod.build(products, campaigns, top=10_000)  # full set into the workbook
    wb = Workbook()

    # Tab 1 — Produkter
    ws = wb.active
    ws.title = "Produkter"
    _header(ws, ["Produkt-ID", "Titel", "Brand", "Product type", "Kampagner",
                 "Forbrug (kr)", "Klik", "Impr", "Konv", "Værdi (kr)", "ROAS", "Note"])
    for p in products:
        pmax = p.get("is_pmax")
        converting = p["conversions_value"] > 0
        note = "PMax (produkt-konv upålidelig)" if pmax else ("konverterer" if converting
                                                              else "forbrug, 0 konv — tal om det")
        ws.append([
            p["product_item_id"], p["product_title"], p["product_brand"], p["product_type_l1"],
            ", ".join(p["campaigns"]), p["cost"], p["clicks"], p["impressions"],
            round(p["conversions"], 1), round(p["conversions_value"], 0), p["roas"], note,
        ])
        fill = GREY if pmax else (GREEN if converting else RED)
        for c in range(1, 13):
            ws.cell(row=ws.max_row, column=c).fill = fill
    _autosize(ws, {1: 12, 2: 40, 3: 14, 4: 14, 5: 30, 6: 12, 7: 8, 8: 8, 9: 8, 10: 12, 11: 8, 12: 30})

    # Tab 2 — Spild (Shopping only, PMax excluded)
    ws2 = wb.create_sheet("Spild")
    _header(ws2, ["Produkt-ID", "Titel", "Kampagner", "Forbrug (kr)", "Klik", "Impr"])
    for p in d["spild_uden_retur"]:
        ws2.append([p["product_item_id"], p["product_title"], ", ".join(p["campaigns"]),
                    p["cost"], p["clicks"], p["impressions"]])
        for c in range(1, 7):
            ws2.cell(row=ws2.max_row, column=c).fill = RED
    _autosize(ws2, {1: 12, 2: 40, 3: 30, 4: 12, 5: 8, 6: 8})

    # Tab 3 — Impression share
    ws3 = wb.create_sheet("ImpressionShare")
    _header(ws3, ["Kampagne", "Kanal", "IS", "Tabt budget", "Tabt rank", "Forbrug (kr)",
                  "Værdi (kr)", "ROAS", "Flaskehals"])
    for c in campaigns:
        bottleneck = "budget" if c["budget_lost_is"] > c["rank_lost_is"] else "rank"
        ws3.append([c["campaign"], c["channel"], round(c["impression_share"], 3),
                    round(c["budget_lost_is"], 3), round(c["rank_lost_is"], 3), c["cost"],
                    round(c["conversions_value"], 0), c["roas"], bottleneck])
    _autosize(ws3, {1: 34, 2: 16, 3: 8, 4: 12, 5: 12, 6: 12, 7: 12, 8: 8, 9: 12})

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Skrev {out} ({len(products)} produkter, {len(d['spild_uden_retur'])} i spild, "
          f"{len(campaigns)} kampagner)")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--products", required=True)
    ap.add_argument("--is", dest="is_file", required=False)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    products = slim_mod.slim_products(json.loads(Path(args.products).read_text()))
    campaigns = slim_mod.slim_is(json.loads(Path(args.is_file).read_text())) if args.is_file else []
    build(products, campaigns, args.out)


if __name__ == "__main__":
    main()
