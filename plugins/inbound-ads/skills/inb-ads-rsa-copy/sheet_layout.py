#!/usr/bin/env python3
"""Single source of truth for the Inbound Google Ads Editor RSA sheet layout.

Both build-template.py (the committed empty template) and fill-sheet.py (the
filled deliverable) call build_sheet() so there is exactly one place that knows
the column order, where the LEN formulas go, and how the red over-length
conditional formatting is applied.

The sheet is one row per RSA: header row 1, then N data rows (rows 2..N+1).
Repeating the same Campaign + Ad Group across rows is how Google Ads Editor
imports several RSAs into the same ad group. Every text column is followed by a
LEN column holding =LEN(<text cell>); a conditional-format rule turns the LEN
cell red when it exceeds the field's hard limit (headline 30, description 90,
path 15).

Because the formulas + formatting live in the xlsx layer (not in CSV values),
they survive upload to Drive and stay live when the client edits the sheet.
"""
import subprocess
import sys


def _ensure_openpyxl():
    """Import openpyxl, pip-installing it if missing, so the skill runs on any
    machine with Python 3 + pip without a manual setup step."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("openpyxl not found - installing...", file=sys.stderr)
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl>=3.1"],
        check=True,
    )


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

RED_FILL = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)

# Field spec: (header, limit-or-None). None = no LEN column (Campaign, Ad Group,
# Ad type, Labels, Final URL, Final mobile URL). A limit means a LEN column with
# a conditional-format rule follows.
#
# "Vinkel" and "Hypotese" are Inbound-internal documentation columns, NOT Google
# Ads Editor fields. They sit LAST, after Final mobile URL. Editor matches import
# columns by name and silently ignores any header it does not recognise, so these
# two are dropped on import and never touch the account. They are deliberately
# named so they cannot collide with a real Editor field (avoid generic names like
# Label / Comment / Status). They document, per RSA row, the overall creative
# angle and the hypothesis behind that ad - and feed inb-ads-rsa-hygiene's
# angle/gap-brief loop later.
FIELDS = (
    [("Campaign", None), ("Ad Group", None), ("Ad type", None), ("Labels", None)]
    + [(f"Headline {i}", 30) for i in range(1, 16)]
    + [(f"Description {i}", 90) for i in range(1, 5)]
    + [("Path 1", 15), ("Path 2", 15)]
    + [("Final URL", None), ("Final mobile URL", None)]
    + [("Vinkel", None), ("Hypotese", None)]
)

# Field -> 1-based text-column index, derived once from FIELDS. LEN columns are
# never addressed by name (fill-sheet only ever writes text cells), so they are
# left out of the map.
COLUMNS: dict[str, int] = {}
_col = 1
for _header, _limit in FIELDS:
    COLUMNS[_header] = _col
    _col += 1
    if _limit is not None:
        _col += 1  # skip the LEN column that follows
LAST_COL = _col - 1


def text_cell(field: str, row: int) -> str:
    """A1 reference of a text field's cell on a given 1-based data row."""
    return f"{get_column_letter(COLUMNS[field])}{row}"


def build_sheet(n_rows: int = 1):
    """Build a workbook with the RSA layout and `n_rows` data rows.

    Returns (workbook, worksheet). Data rows are rows 2..n_rows+1, each pre-wired
    with LEN formulas and the red over-length conditional formatting. Only the
    constant "Ad type" cell is pre-seeded; the caller fills the rest.
    """
    if n_rows < 1:
        raise ValueError("n_rows must be >= 1")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RSA"

    col = 1  # 1-based column index
    for header, limit in FIELDS:
        text_col = col
        text_letter = get_column_letter(text_col)
        hc = ws.cell(row=1, column=text_col, value=header)
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = Alignment(horizontal="center")
        col += 1

        if limit is not None:
            len_letter = get_column_letter(col)
            lc = ws.cell(row=1, column=col, value="LEN")
            lc.fill = HEADER_FILL
            lc.font = HEADER_FONT
            lc.alignment = Alignment(horizontal="center")
            # Live length formula per data row.
            for r in range(2, n_rows + 2):
                ws[f"{len_letter}{r}"] = f"=LEN({text_letter}{r})"
            # One red-over-limit rule covering the whole LEN column range.
            rng = f"{len_letter}2:{len_letter}{n_rows + 1}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThan", formula=[str(limit)], fill=RED_FILL),
            )
            col += 1

    # Pre-seed the constant value on every data row. Campaign/Ad Group are
    # written per run by fill-sheet.py.
    for r in range(2, n_rows + 2):
        ws.cell(row=r, column=COLUMNS["Ad type"], value="Responsive search ad")

    # Sensible default widths; fill-sheet.py auto-resizes after writing text.
    for c in range(1, LAST_COL + 1):
        letter = get_column_letter(c)
        header = ws.cell(row=1, column=c).value
        ws.column_dimensions[letter].width = 6 if header == "LEN" else 18

    return wb, ws


def autosize_columns(ws, min_width: int = 6, max_width: int = 60, padding: int = 2) -> None:
    """Set each column's width to fit its widest cell (header or data).

    LEN cells hold =LEN(...) formulas that would measure as the formula string
    instead of the rendered number, so any LEN-header column gets a fixed narrow
    width and formula cells are skipped elsewhere.
    """
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        header = col_cells[0].value
        letter = get_column_letter(col_idx)
        if header == "LEN":
            ws.column_dimensions[letter].width = 6
            continue
        widest = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            text = str(v)
            if text.startswith("="):
                continue
            widest = max(widest, len(text))
        width = max(min_width, min(max_width, widest + padding))
        ws.column_dimensions[letter].width = width
