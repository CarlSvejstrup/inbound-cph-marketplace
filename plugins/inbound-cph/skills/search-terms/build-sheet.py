#!/usr/bin/env python3
"""Build the search-terms analysis workbook (.xlsx) for the search-terms skill.

Layout follows the field-tested template a user produced for Dansk Studie Center,
with one addition: a dedicated Vindere tab (converting terms not yet exact keywords).
Eight sheets:
  Oversigt | Alle search terms | Godt placeret (ingen handling) | Vindere (promover til exact)
  | Placement-problem | Irrelevante (tilfoej negativ) | Graensetilfaelde | Anbefalede negatives

The Placement-problem tab has three extra columns (Placement-aarsag, Ad-group LP,
Top annonce-tema) that the other term tabs do not - intent context lives there.

Colours, header style, classification row-fills, and freeze panes are baked into the
.xlsx layer, so they survive upload to Drive and render when opened in Google Sheets.
No gws / Sheets API. Runs in Cowork and locally.

Classification values (used to colour the Klassificering column):
  RELEVANT          -> green  C6EFCE
  VINDER            -> teal   A9D08E (distinct from RELEVANT)
  PLACEMENT_PROBLEM -> yellow FFEB9C  (sub-typed via row's placement_reason: struktur | intent)
  IRRELEVANT        -> red    FFC7CE
  GRAENSE           -> blue   D9E1F2

Input JSON schema (lists may be empty; missing keys render blank):
{
  "client": "Dansk Studie Center",
  "account_id": "3069826320",
  "period": "13. feb 2026 - 12. maj 2026 (3 mdr.)",
  "scope": "Kun aktive Search-kampagner",
  "filter": "Search terms med >5 DKK spend",
  "offering": ["Grupperejser for unge ...", "Destinationer: ...", ...],  // landing-page grounded
  "method_notes": ["Data hentet via GAQL ...", "Cost konverteret fra micros ...", ...],
  "distribution": [                       // Oversigt fordelingstabel; order as you want shown
     {"kategori":"GODT PLACERET (ingen handling)","antal":123,"spend":10419.44},
     {"kategori":"VINDER (promover til exact)","antal":12,"spend":4200.0},
     ...
  ],
  // One list per classification tab. Each row is a full term record:
  "rows": [
    {"term","match_type","kampagne","ad_group","keyword","keyword_match_type",
     "spend","impr","clicks","ctr","conv","klassificering","begrundelse",
     // PLACEMENT_PROBLEM rows only (ignored on other buckets):
     "placement_reason",  // "struktur" | "intent"
     "ad_group_lp",       // the served ad group's final URL (from ad_group_ad pull)
     "ad_group_top_theme" // 1-line synthesis of the served ad group's top headlines
    }
  ],
  // Anbefalede negatives: synthesised import-ready list.
  "negatives": [
    {"negative","match_type","hvor","spildt_budget","begrundelse"}
  ]
}

`rows` holds every term once (aggregated per term, with keyword/match listed). The skill
splits them into the per-tab views by `klassificering`; this script does the routing.

Run:
  python3 build-sheet.py --in analysis.json --out "Search Terms - <klient> - <dato>.xlsx"
"""
import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Palette (matches the user's template exactly, plus VINDER).
HEADER_BG = "1F4E78"
HEADER_FG = "FFFFFF"
FILL = {
    "RELEVANT": "C6EFCE",
    "VINDER": "A9D08E",
    "PLACEMENT_PROBLEM": "FFEB9C",
    "IRRELEVANT": "FFC7CE",
    "GRAENSE": "D9E1F2",
}

HEADER_FONT = Font(bold=True, color=HEADER_FG)
TITLE_FONT = Font(bold=True, size=16)
SECTION_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

TERM_HEADERS = [
    "Search term", "Match type", "Kampagne", "Ad group", "Triggerende keyword",
    "Keyword match type", "Budget brugt (DKK)", "Impressions", "Klik", "CTR (%)",
    "Konverteringer", "Klassificering", "Begrundelse / Anbefaling",
]
TERM_KEYS = [
    "term", "match_type", "kampagne", "ad_group", "keyword", "keyword_match_type",
    "spend", "impr", "clicks", "ctr", "conv", "klassificering", "begrundelse",
]
NEG_HEADERS = [
    "Negative keyword", "Anbefalet match type", "Hvor (kampagne/konto-niveau)",
    "Spildt budget (DKK) sidste periode", "Begrundelse",
]
NEG_KEYS = ["negative", "match_type", "hvor", "spildt_budget", "begrundelse"]

# Column widths for the term tabs (index-aligned to TERM_HEADERS).
TERM_WIDTHS = [32, 12, 26, 22, 24, 16, 16, 12, 8, 9, 13, 18, 60]
NEG_WIDTHS = [28, 18, 36, 22, 60]

# Extra columns appended on the Placement-problem tab only (intent context).
PLACEMENT_EXTRA_HEADERS = ["Placement-aarsag", "Ad-group LP", "Top annonce-tema"]
PLACEMENT_EXTRA_KEYS = ["placement_reason", "ad_group_lp", "ad_group_top_theme"]
PLACEMENT_EXTRA_WIDTHS = [16, 44, 40]


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _round(v):
    return round(v, 2) if isinstance(v, (int, float)) else v


def _header_row(ws, headers, widths):
    fill = _fill(HEADER_BG)
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_term_tab(wb, title, rows, extra=False):
    """Write a term tab. If extra=True (only for the placement tab), append the
    three intent-context columns: Placement-aarsag, Ad-group LP, Top annonce-tema."""
    ws = wb.create_sheet(title)
    headers = TERM_HEADERS + (PLACEMENT_EXTRA_HEADERS if extra else [])
    keys = TERM_KEYS + (PLACEMENT_EXTRA_KEYS if extra else [])
    widths = TERM_WIDTHS + (PLACEMENT_EXTRA_WIDTHS if extra else [])
    _header_row(ws, headers, widths)
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(keys, start=1):
            val = row.get(key)
            if key in ("spend", "ctr"):
                val = _round(val)
            ws.cell(row=r, column=c, value=val).alignment = WRAP
        # Colour the Klassificering cell (col 12) by bucket.
        klass = row.get("klassificering")
        if klass in FILL:
            ws.cell(row=r, column=12).fill = _fill(FILL[klass])
    return ws


def build(data, out_path):
    wb = openpyxl.Workbook()
    rows = data.get("rows", [])

    def by_class(*names):
        s = set(names)
        return [r for r in rows if r.get("klassificering") in s]

    # --- Oversigt ---
    ws = wb.active
    ws.title = "Oversigt"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws["A1"] = f"Search Terms Analyse - {data.get('client', '')}"
    ws["A1"].font = TITLE_FONT
    meta = [
        ("Konto-ID:", data.get("account_id", "")),
        ("Periode:", data.get("period", "")),
        ("Scope:", data.get("scope", "Kun aktive Search-kampagner")),
        ("Filter:", data.get("filter", "")),
    ]
    r = 2
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Fordeling").font = SECTION_FONT
    r += 1
    for c, h in enumerate(["Kategori", "Antal", "Spend (DKK)"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill(HEADER_BG)
        cell.font = HEADER_FONT
    r += 1
    first_dist = r
    for d in data.get("distribution", []):
        ws.cell(row=r, column=1, value=d.get("kategori"))
        ws.cell(row=r, column=2, value=d.get("antal"))
        ws.cell(row=r, column=3, value=_round(d.get("spend")))
        # tint the kategori cell by leading keyword
        k = (d.get("kategori") or "").upper()
        for name, hexc in FILL.items():
            if name.split("_")[0] in k or (name == "GRAENSE" and "GR" in k):
                ws.cell(row=r, column=1).fill = _fill(hexc)
                break
        r += 1
    last_dist = r - 1
    if last_dist >= first_dist:
        tot = ws.cell(row=r, column=1, value="TOTAL")
        tot.font = SECTION_FONT
        ws.cell(row=r, column=2, value=f"=SUM(B{first_dist}:B{last_dist})").font = SECTION_FONT
        ws.cell(row=r, column=3, value=f"=SUM(C{first_dist}:C{last_dist})").font = SECTION_FONT
    r += 2

    # Klientens udbud (landing-page grounded)
    offering = data.get("offering", [])
    if offering:
        ws.cell(row=r, column=1, value="Klientens udbud (fra landingsside)").font = SECTION_FONT
        r += 1
        for line in offering:
            ws.cell(row=r, column=1, value=line).alignment = WRAP
            r += 1
        r += 1

    notes = data.get("method_notes", [])
    if notes:
        ws.cell(row=r, column=1, value="Metode & noter").font = SECTION_FONT
        r += 1
        for line in notes:
            ws.cell(row=r, column=1, value=line).alignment = WRAP
            r += 1

    # --- Alle search terms (every row) ---
    _write_term_tab(wb, "Alle search terms", rows)

    # --- Per-class tabs ---
    _write_term_tab(wb, "Godt placeret (ingen handling)", by_class("RELEVANT"))
    _write_term_tab(wb, "Vindere (promover til exact)", by_class("VINDER"))
    _write_term_tab(wb, "Placement-problem", by_class("PLACEMENT_PROBLEM"), extra=True)
    _write_term_tab(wb, "Irrelevante (tilfoej negativ)", by_class("IRRELEVANT"))
    _write_term_tab(wb, "Graensetilfaelde", by_class("GRAENSE"))

    # --- Anbefalede negatives (import-ready) ---
    ws = wb.create_sheet("Anbefalede negatives")
    _header_row(ws, NEG_HEADERS, NEG_WIDTHS)
    for r2, neg in enumerate(data.get("negatives", []), start=2):
        for c, key in enumerate(NEG_KEYS, start=1):
            val = neg.get(key)
            if key == "spildt_budget":
                val = _round(val)
            ws.cell(row=r2, column=c, value=val).alignment = WRAP

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="analysis JSON path")
    ap.add_argument("--out", dest="out", required=True, help="output .xlsx path")
    args = ap.parse_args()
    data = json.loads(Path(args.inp).read_text(encoding="utf-8"))
    print(build(data, args.out))


if __name__ == "__main__":
    main()
