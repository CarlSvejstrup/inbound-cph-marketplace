#!/usr/bin/env python3
"""Build the annonce-optimering diagnostic workbook (.xlsx).

This skill is a POST-LAUNCH ASSET-HYGIENE diagnostic, not a profit classifier.
It was deliberately reshaped after a live test against real Inbound accounts
(2026-05-29) proved two things:

  1. Google's per-asset performance_label comes back NOT_APPLICABLE / PENDING on
     Inbound's accounts (too low volume) - never BEST/GOOD/LOW. So the label
     cannot be the primary signal.
  2. Per-asset CTR/CVR from ad_group_ad_asset_view is confounded: an RSA serves
     ~3 headlines + 2 descriptions per impression and the same click/conversion
     is attributed to EVERY served asset. Per-asset CVR therefore cannot carry a
     four-quadrant Winners/Hidden Gems/Money Pits/Losers matrix, and the
     conversion counts (0/1/2 per asset) are far below significance anyway.

So this workbook reports STRUCTURAL FACTS that hold without significance:
  - champion-challenger coverage (RSAs per ad group; <2 = build a challenger),
  - dead-weight assets (never-served / near-zero impressions = a coverage fact),
  - angle-coverage gaps per ad group (which angles have no served asset),
  - Google's label ONLY when it is BEST/GOOD/LOW (else "ikke nok data endnu"),
  - an optional CVR hint gated behind a hard significance floor (else
    "utilstraekkelig data").

The angle-gap output is the gap-brief fed back into annoncetekster-v2 to write
the next challenger headlines - that is where the build->operate->iterate loop
closes.

Colours, header style, freeze panes are baked into the .xlsx layer so they
survive upload to Drive and render when opened in Google Sheets. No gws /
Sheets API. Runs in Cowork and locally.

Input JSON schema (lists may be empty; missing keys render blank):
{
  "client": "Dansk Studie Center",
  "account_id": "3069826320",
  "period": "1. maj 2026 - 29. maj 2026",
  "scope": "Kun aktive Search-kampagner og aktive RSA'er",
  "min_impressions": 50,                 // the dead-weight / significance floor, surfaced in Oversigt
  "method_notes": ["Data via GAQL ad_group_ad_asset_view ...", ...],

  // Per ad group: RSA count + coverage flag (champion-challenger).
  "ad_groups": [
    {"kampagne": "...", "ad_group": "...", "rsa_count": 1,
     "challenger_flag": true,            // true when rsa_count < 2
     "manglende_vinkler": ["urgency", "CTA"]}   // angles with no served asset -> gap-brief
  ],

  // Per asset row (headline/description). Grouped into ONE TAB PER AD GROUP by
  // (kampagne, ad_group). The Google-label and CVR columns were intentionally
  // dropped from the sheet: on Inbound's accounts every row is identical
  // ("ikke nok data endnu" / "utilstraekkelig data"), so they were noise. The
  // skill may still compute cvr_hint internally for the significance gate, but
  // it is not rendered.
  "assets": [
    {"kampagne","ad_group","felt_type","tekst","vinkel",
     "impressions","clicks","cost",
     "status",                // "DOEDVAEGT" | "AKTIV" | "FOR_NY" (shown as DØDVÆGT / AKTIV / FOR NY)
     "anbefaling"}            // recommend-only text
  ],

  // Synthesised gap-brief for annoncetekster-v2 (angles missing a served asset, per ad group).
  "gap_brief": [
    {"kampagne","ad_group","manglende_vinkler","forslag"}
  ]
}

Run:
  python3 build-sheet.py --in analysis.json --out "Annonce-optimering - <klient> - <dato>.xlsx"
"""
import argparse
import json
import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    """Import openpyxl, pip-installing it if missing, so this runs on any
    machine with Python 3 + pip (same self-bootstrap as the annoncetekster
    scripts)."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("openpyxl not found - installing...", file=sys.stderr)
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl>=3.1"],
        check=True,
    )


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

HEADER_BG = "1F4E78"
HEADER_FG = "FFFFFF"
# Status colours: dead weight = red, active = green, too-new = blue.
# Keys are the ASCII enums the skill writes into status; STATUS_DISPLAY maps
# them to the real Danish shown in the sheet.
STATUS_FILL = {
    "DOEDVAEGT": "FFC7CE",
    "AKTIV": "C6EFCE",
    "FOR_NY": "D9E1F2",
}
STATUS_DISPLAY = {
    "DOEDVAEGT": "DØDVÆGT",
    "AKTIV": "AKTIV",
    "FOR_NY": "FOR NY",
}
FLAG_FILL = "FFEB9C"  # yellow for challenger flags / missing angles

HEADER_FONT = Font(bold=True, color=HEADER_FG)
TITLE_FONT = Font(bold=True, size=16)
SECTION_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

AG_HEADERS = [
    "Kampagne", "Ad group", "Antal aktive RSA", "Byg challenger?",
    "Manglende vinkler (gap-brief)",
]
AG_WIDTHS = [26, 24, 16, 16, 40]

# Per-ad-group asset tab. Google-label and CVR-indikation columns were removed:
# on Inbound's accounts they are always identical ("ikke nok data endnu" /
# "utilstrækkelig data"), so they carried no information.
ASSET_HEADERS = [
    "Felt", "Tekst", "Vinkel", "Impressions", "Klik", "Spend (DKK)",
    "Status", "Anbefaling",
]
ASSET_KEYS = [
    "felt_type", "tekst", "vinkel", "impressions", "clicks", "cost",
    "status", "anbefaling",
]
ASSET_WIDTHS = [12, 42, 16, 12, 8, 12, 12, 52]
# 0-based index of the Status column within ASSET_KEYS (for colouring).
STATUS_COL = ASSET_KEYS.index("status") + 1

GAP_HEADERS = ["Kampagne", "Ad group", "Manglende vinkler", "Forslag til challenger"]
GAP_KEYS = ["kampagne", "ad_group", "manglende_vinkler", "forslag"]
GAP_WIDTHS = [24, 24, 30, 56]


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _round(v):
    return round(v, 2) if isinstance(v, (int, float)) else v


def _join(v):
    return ", ".join(v) if isinstance(v, (list, tuple)) else v


def _header_row(ws, headers, widths):
    fill = _fill(HEADER_BG)
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _oversigt(wb, data):
    ws = wb.active
    ws.title = "Oversigt"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
    ws["A1"] = f"Annonce-optimering - {data.get('client', '')}"
    ws["A1"].font = TITLE_FONT
    meta = [
        ("Konto-ID:", data.get("account_id", "")),
        ("Periode:", data.get("period", "")),
        ("Scope:", data.get("scope", "Kun aktive Search-kampagner og aktive RSA'er")),
        ("Sample-floor (min. impressions):", data.get("min_impressions", "")),
    ]
    r = 2
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1
    # What this report is / is NOT - honesty banner so no one mis-reads it.
    ws.cell(row=r, column=1, value="Hvad rapporten er").font = SECTION_FONT
    r += 1
    for line in [
        "Strukturel asset-hygiejne: RSA-dækning, dødvægt-assets, vinkel-huller.",
        "Den dømmer IKKE assets på konverteringsrate - per-asset CVR i Google er",
        "konfunderet (samme klik tilskrives alle serverede assets) og under signifikans.",
        "Alt er anbefalinger - intet redigeres, pauses eller skrives til kontoen.",
    ]:
        ws.cell(row=r, column=1, value=line).alignment = WRAP
        r += 1
    r += 1
    notes = data.get("method_notes", [])
    if notes:
        ws.cell(row=r, column=1, value="Metode & noter").font = SECTION_FONT
        r += 1
        for line in notes:
            ws.cell(row=r, column=1, value=line).alignment = WRAP
            r += 1


def _ad_group_tab(wb, data):
    ws = wb.create_sheet("Ad group-dækning")
    _header_row(ws, AG_HEADERS, AG_WIDTHS)
    for r, row in enumerate(data.get("ad_groups", []), start=2):
        ws.cell(row=r, column=1, value=row.get("kampagne")).alignment = WRAP
        ws.cell(row=r, column=2, value=row.get("ad_group")).alignment = WRAP
        ws.cell(row=r, column=3, value=row.get("rsa_count"))
        flag = bool(row.get("challenger_flag"))
        fc = ws.cell(row=r, column=4, value="JA - byg challenger" if flag else "OK")  # noqa
        if flag:
            fc.fill = _fill(FLAG_FILL)
        mc = ws.cell(row=r, column=5, value=_join(row.get("manglende_vinkler")))
        mc.alignment = WRAP
        if row.get("manglende_vinkler"):
            mc.fill = _fill(FLAG_FILL)


_TAB_FORBIDDEN = ':\\/?*[]'


def _safe_tab_name(base, used):
    """Excel tab names: max 31 chars, none of : \\ / ? * [ ], and unique.
    Sanitise, truncate, and de-dupe with a numeric suffix."""
    name = "".join("-" if ch in _TAB_FORBIDDEN else ch for ch in (base or "")).strip()
    name = name or "Ad group"
    name = name[:31]
    if name not in used:
        used.add(name)
        return name
    # De-dupe: append " (n)" while staying within 31 chars.
    i = 2
    while True:
        suffix = f" ({i})"
        candidate = name[:31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def _ad_group_overview(ws, start_row, kampagne, ad_group, group_assets, ag_meta):
    """Write a compact per-ad-group overview block. Returns the row after it.

    Shows the facts that hold without significance: campaign + ad group, RSA
    count + challenger flag, asset counts split by status and by field type, and
    the missing angles (gap-brief). All derived from data we already have."""
    r = start_row
    ws.cell(row=r, column=1, value="Ad group:").font = SECTION_FONT
    ws.cell(row=r, column=2, value=ad_group)
    r += 1
    ws.cell(row=r, column=1, value="Kampagne:").font = SECTION_FONT
    ws.cell(row=r, column=2, value=kampagne)
    r += 1

    ws.cell(row=r, column=1, value="Overblik").font = SECTION_FONT
    r += 1

    # Counts derived from this group's assets.
    n_assets = len(group_assets)
    by_status = {"AKTIV": 0, "DOEDVAEGT": 0, "FOR_NY": 0}
    n_head = n_desc = 0
    for a in group_assets:
        st = a.get("status")
        if st in by_status:
            by_status[st] += 1
        ft = (a.get("felt_type") or "").upper()
        if ft == "HEADLINE":
            n_head += 1
        elif ft == "DESCRIPTION":
            n_desc += 1

    rsa_count = ag_meta.get("rsa_count")
    challenger = bool(ag_meta.get("challenger_flag"))
    mangler = ag_meta.get("manglende_vinkler") or []

    rows = [
        ("Aktive RSA i gruppen:", rsa_count if rsa_count is not None else "?"),
        ("Byg challenger?", "JA - kun under 2 RSA" if challenger else "OK - mindst 2 RSA"),
        ("Assets i alt:", f"{n_assets} ({n_head} headlines, {n_desc} descriptions)"),
        ("Status-fordeling:",
         f"{by_status['AKTIV']} aktive / {by_status['DOEDVAEGT']} dødvægt / {by_status['FOR_NY']} for ny"),
        ("Manglende vinkler:", ", ".join(mangler) if mangler else "ingen - fuldt dækket"),
    ]
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.alignment = WRAP
        # Highlight the two action-bearing lines.
        if label == "Byg challenger?" and challenger:
            c.fill = _fill(FLAG_FILL)
        if label == "Manglende vinkler:" and mangler:
            c.fill = _fill(FLAG_FILL)
        r += 1
    return r + 1  # blank spacer row after the block


def _asset_tabs(wb, data):
    """One tab per ad group. Each tab opens with a per-ad-group overview block,
    then the asset table for that group."""
    assets = data.get("assets", [])
    # Group assets by (kampagne, ad_group), preserving first-seen order.
    groups = {}
    order = []
    for row in assets:
        key = (row.get("kampagne", ""), row.get("ad_group", ""))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)

    # Lookup of ad-group metadata (rsa_count, challenger_flag, manglende_vinkler).
    ag_lookup = {
        (g.get("kampagne", ""), g.get("ad_group", "")): g
        for g in data.get("ad_groups", [])
    }

    used_names = set()
    for kampagne, ad_group in order:
        ws = wb.create_sheet(_safe_tab_name(ad_group, used_names))
        ag_meta = ag_lookup.get((kampagne, ad_group), {})
        # Overview block (full names live here; the tab label may be truncated).
        head_row = _ad_group_overview(ws, 1, kampagne, ad_group, groups[(kampagne, ad_group)], ag_meta)
        fill = _fill(HEADER_BG)
        for c, head in enumerate(ASSET_HEADERS, start=1):
            cell = ws.cell(row=head_row, column=c, value=head)
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = WRAP
        for i, w in enumerate(ASSET_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{head_row + 1}"
        for r, row in enumerate(groups[(kampagne, ad_group)], start=head_row + 1):
            for c, key in enumerate(ASSET_KEYS, start=1):
                val = row.get(key)
                if key == "cost":
                    val = _round(val)
                if key == "status":
                    val = STATUS_DISPLAY.get(val, val)
                ws.cell(row=r, column=c, value=val).alignment = WRAP
            status = row.get("status")
            if status in STATUS_FILL:
                ws.cell(row=r, column=STATUS_COL).fill = _fill(STATUS_FILL[status])


def _gap_tab(wb, data):
    ws = wb.create_sheet("Gap-brief")
    _header_row(ws, GAP_HEADERS, GAP_WIDTHS)
    for r, row in enumerate(data.get("gap_brief", []), start=2):
        ws.cell(row=r, column=1, value=row.get("kampagne")).alignment = WRAP
        ws.cell(row=r, column=2, value=row.get("ad_group")).alignment = WRAP
        ws.cell(row=r, column=3, value=_join(row.get("manglende_vinkler"))).alignment = WRAP
        ws.cell(row=r, column=4, value=row.get("forslag")).alignment = WRAP


def build(data, out_path):
    wb = openpyxl.Workbook()
    _oversigt(wb, data)
    _ad_group_tab(wb, data)
    _asset_tabs(wb, data)
    _gap_tab(wb, data)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="analysis JSON path")
    ap.add_argument("--out", dest="out", required=True, help="output .xlsx path")
    args = ap.parse_args()
    data = json.loads(Path(args.inp).read_text(encoding="utf-8"))
    print(build(data, args.out))


if __name__ == "__main__":
    main()
