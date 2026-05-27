#!/usr/bin/env python3
"""Generate template.xlsx for the annoncetekster skill.

Builds Inbound's Google Ads Editor RSA import layout as a real .xlsx so that
=LEN() formulas and red over-length conditional formatting are baked into the
file. The skill fills a copy of this template and saves a new file each run;
because formulas + formatting live in the xlsx layer (not in CSV values), they
survive upload to Drive and stay live when the client edits the sheet.

Layout (row 1 = headers, row 2 = single data row):
  Campaign | Ad Group | Ad type | Labels |
  Headline 1 | LEN | ... | Headline 15 | LEN |
  Description 1 | LEN | ... | Description 4 | LEN |
  Path 1 | LEN | Path 2 | LEN |
  Final URL | Final mobile URL

Every text column is followed by a LEN column holding =LEN(<text cell>).
Conditional formatting turns a LEN cell red when it exceeds the field's limit
(headline 30, description 90, path 15).

Run: python3 build-template.py   ->  writes template.xlsx next to this script.
Deterministic: re-running reproduces the same file.
"""
import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    """Import openpyxl, pip-installing it if missing, so this runs on any
    machine with Python 3 + pip."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("openpyxl not found - installing...", file=sys.stderr)
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl>=3.1"],
        check=True,
    )


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

RED_FILL = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)

# Field spec: (header, limit-or-None). None = no LEN column (Campaign, Ad Group,
# Ad type, Labels, Final URL, Final mobile URL). A limit means a LEN column with
# a conditional-format rule follows.
FIELDS = (
    [("Campaign", None), ("Ad Group", None), ("Ad type", None), ("Labels", None)]
    + [(f"Headline {i}", 30) for i in range(1, 16)]
    + [(f"Description {i}", 90) for i in range(1, 5)]
    + [("Path 1", 15), ("Path 2", 15)]
    + [("Final URL", None), ("Final mobile URL", None)]
)


def build() -> dict:
    """Build the workbook and return the field -> data-cell A1 map."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RSA"

    col = 1  # 1-based column index
    cell_map: dict[str, str] = {}

    for header, limit in FIELDS:
        text_col = col
        text_letter = get_column_letter(text_col)
        # Header
        hc = ws.cell(row=1, column=text_col, value=header)
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = Alignment(horizontal="center")
        cell_map[header] = f"{text_letter}2"
        col += 1

        if limit is not None:
            len_letter = get_column_letter(col)
            lc = ws.cell(row=1, column=col, value="LEN")
            lc.fill = HEADER_FILL
            lc.font = HEADER_FONT
            lc.alignment = Alignment(horizontal="center")
            # Live length formula in the data row
            ws[f"{len_letter}2"] = f"=LEN({text_letter}2)"
            # Red when over the field's hard limit
            ws.conditional_formatting.add(
                f"{len_letter}2",
                CellIsRule(operator="greaterThan", formula=[str(limit)], fill=RED_FILL),
            )
            col += 1

    # Pre-seed only the constant value. Campaign is built per-run from
    # AskUserQuestion intake; fill-sheet.py overwrites cell A2.
    ws[cell_map["Ad type"]] = "Responsive search ad"

    # Column widths: narrow for LEN cells, wider for text. fill-sheet.py
    # auto-resizes again after writing the actual copy, so these are just
    # sensible defaults for an empty template.
    for c in range(1, col):
        letter = get_column_letter(c)
        header = ws.cell(row=1, column=c).value
        ws.column_dimensions[letter].width = 6 if header == "LEN" else 18

    out = Path(__file__).with_name("template.xlsx")
    wb.save(out)
    return cell_map


if __name__ == "__main__":
    cmap = build()
    print("Wrote template.xlsx")
    print("Data-cell map (field -> A1, row 2):")
    for field, ref in cmap.items():
        print(f"  {field:18s} {ref}")
