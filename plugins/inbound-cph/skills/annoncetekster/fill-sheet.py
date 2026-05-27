#!/usr/bin/env python3
"""Fill template.xlsx with generated RSA copy and save a new file.

The skill calls this after it has generated and length-checked the copy. It
loads the bundled template (which already holds the =LEN() formulas and the red
over-length conditional formatting), writes only the text cells (never the LEN
cells), and saves a fresh .xlsx. The formulas + formatting ride along untouched.

Usage:
  python3 fill-sheet.py --copy copy.json --out "/path/RSA - Client - 2026-05-27.xlsx"

copy.json shape (all keys optional except headlines/descriptions):
  {
    "campaign": "IC | GSN | Generic |",   # optional, default kept from template
    "ad_group": "",                          # optional
    "headlines": ["...", ...],               # up to 15, each <= 30 chars
    "descriptions": ["...", ...],            # up to 4, each <= 90 chars
    "paths": ["...", "..."],                 # up to 2, each <= 15 chars
    "final_url": "https://...",
    "final_mobile_url": ""                   # optional
  }

Exits non-zero and prints the offending fields if any string exceeds its hard
limit, so the skill never writes an over-length sheet.
"""
import argparse
import json
import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    """Import openpyxl, pip-installing it if missing. Lets the skill run on any
    machine with Python 3 + pip without a manual setup step."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("openpyxl not found - installing...", file=sys.stderr)
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl>=3.1"],
        check=True,
    )


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

TEMPLATE = Path(__file__).with_name("template.xlsx")

HEADLINE_CELLS = ["E2", "G2", "I2", "K2", "M2", "O2", "Q2", "S2", "U2", "W2",
                  "Y2", "AA2", "AC2", "AE2", "AG2"]
DESCRIPTION_CELLS = ["AI2", "AK2", "AM2", "AO2"]
PATH_CELLS = ["AQ2", "AS2"]
CAMPAIGN_CELL = "A2"
AD_GROUP_CELL = "B2"
FINAL_URL_CELL = "AU2"
FINAL_MOBILE_URL_CELL = "AV2"

LIMITS = {"headline": 30, "description": 90, "path": 15}


def validate(copy: dict) -> list[str]:
    errs = []
    for i, h in enumerate(copy.get("headlines", []), 1):
        if len(h) > LIMITS["headline"]:
            errs.append(f"Headline {i} ({len(h)} > 30): {h!r}")
    for i, d in enumerate(copy.get("descriptions", []), 1):
        if len(d) > LIMITS["description"]:
            errs.append(f"Description {i} ({len(d)} > 90): {d!r}")
    for i, p in enumerate(copy.get("paths", []), 1):
        if len(p) > LIMITS["path"]:
            errs.append(f"Path {i} ({len(p)} > 15): {p!r}")
    return errs


def _autosize_columns(ws, min_width: int = 6, max_width: int = 60, padding: int = 2) -> None:
    """Set each column's width to fit its widest cell (header or data).

    LEN cells holding =LEN(...) formulas would be measured as the formula string
    (long) instead of the rendered number (short). We skip those by giving any
    LEN-header column a fixed narrow width.
    """
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        header = col_cells[0].value
        letter = get_column_letter(col_idx)
        if header == "LEN":
            ws.column_dimensions[letter].width = 6
            continue
        widest = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            text = str(v)
            if text.startswith("="):
                continue
            widest = max(widest, len(text))
        width = max(min_width, min(max_width, widest + padding))
        ws.column_dimensions[letter].width = width


def fill(copy: dict, out: Path) -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["RSA"]

    if copy.get("campaign"):
        ws[CAMPAIGN_CELL] = copy["campaign"]
    if copy.get("ad_group"):
        ws[AD_GROUP_CELL] = copy["ad_group"]

    for cell, value in zip(HEADLINE_CELLS, copy.get("headlines", [])):
        ws[cell] = value
    for cell, value in zip(DESCRIPTION_CELLS, copy.get("descriptions", [])):
        ws[cell] = value
    for cell, value in zip(PATH_CELLS, copy.get("paths", [])):
        ws[cell] = value
    if copy.get("final_url"):
        ws[FINAL_URL_CELL] = copy["final_url"]
    if copy.get("final_mobile_url"):
        ws[FINAL_MOBILE_URL_CELL] = copy["final_mobile_url"]

    _autosize_columns(ws)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--copy", required=True, help="path to copy.json")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    copy = json.loads(Path(args.copy).read_text())
    errs = validate(copy)
    if errs:
        print("REFUSING TO WRITE - over-length fields:", file=sys.stderr)
        for e in errs:
            print("  " + e, file=sys.stderr)
        return 1

    out = Path(args.out)
    fill(copy, out)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
