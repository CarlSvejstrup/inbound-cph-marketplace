#!/usr/bin/env python3
"""Phase-4 assembler for campaign-build.

Pure transform: merges the four upstream JSON shapes (campaign-strategy, structuring,
rsa manifest, assets) into Ian's 10-tab review workbook, then runs validation. NO Google
Ads API push, no external reads/writes — it only reads local JSON and writes one .xlsx.

The assembler is Excel-ONLY (decision 2026-06-05). The workbook is the client-confirmation
artifact; Editor CSVs are generated LATER, from the confirmed Excel, by the separate
`google-ads-general` converter skill. The workbook is therefore a lossless superset — every
field a CSV needs has a dedicated workbook cell (Max CPC, numeric daily budget, negative
level/ad-group, per-type asset columns). The CSV = Editor-schema-only boundary now lives in
the converter, not here.

Reuses the hard field limits (headline 30 / description 90 / path 15) and the LEN+red
conditional-formatting technique from responsive-search-ads/sheet_layout.py — there is one
source of truth for those, not a retyped copy here.

Two hard emit-time guards (defense-in-depth, see references/assembler-contract.md):
  1. Every positive keyword row must have an explicit Exact or Phrase match type. A blank /
     missing / Broad match type makes Editor create a Broad keyword, violating the locked
     "Exact + selected Phrase, no Broad" policy. The assembler REFUSES to emit keywords.csv
     if any row fails this.
  2. Tab 09 validation recomputes LEN + Pass independently against the imported limits, so a
     human edit between Phase 3 and assembly that pushes a field over-length is still caught.

The inherited 277-term MCC shared negative list is NEVER enumerated — it is applied by
reference (a tab-08 launch-gate line + a single reference line in tab 04). Only client-specific
additions become negative rows. Monitor-first candidates go to tab 05 only.
"""
import argparse
import importlib.util
import json
import os
import sys

# --- import the RSA layout module for the limits + CF technique (single source of truth) ---
_HERE = os.path.dirname(os.path.abspath(__file__))
_RSA_LAYOUT = os.path.normpath(
    os.path.join(_HERE, "..", "responsive-search-ads", "sheet_layout.py")
)


def _load_rsa_layout():
    spec = importlib.util.spec_from_file_location("rsa_sheet_layout", _RSA_LAYOUT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_layout = _load_rsa_layout()  # also pip-installs openpyxl on first run via its bootstrap
import openpyxl  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# Hard limits derived from the RSA layout's FIELDS — never retyped. FIELDS entries are
# (header, limit-or-None); the limit on "Headline 1"/"Description 1"/"Path 1" is the cap.
_LIMITS = {h.split()[0]: lim for h, lim in _layout.FIELDS if lim is not None}
HEADLINE_LIMIT = _LIMITS["Headline"]      # 30
DESCRIPTION_LIMIT = _LIMITS["Description"]  # 90
PATH_LIMIT = _LIMITS["Path"]               # 15

SHARED_NEG_NAME = "Generelle negative søgeord"
SHARED_NEG_ID = "6688642473"
SHARED_NEG_MCC = "1138360630"

# --- Visual design system (client-facing: the workbook is often sent to the customer) -------
# Inbound dark-navy header, white bold text, soft zebra banding, thin grid. Row 1 STAYS the
# header row (the google-ads-general converter reads headers from row 1) — we style cells only,
# never insert a banner row above the data. Styling must not move the header off row 1.
NAVY = "1F2A44"          # Inbound dark navy
NAVY_TEXT = "FFFFFF"
BAND_FILL = "F4F6FA"     # very light blue-grey for odd data rows (zebra)
RED_FILL_HEX = "F4C7C3"  # over-length flag (kept)
GREEN_TEXT = "1E7A34"    # Pass=True
RED_TEXT = "B00020"      # Pass=False

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BAND_PATTERN = PatternFill(start_color=BAND_FILL, end_color=BAND_FILL, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_FILL_HEX, end_color=RED_FILL_HEX, fill_type="solid")
HEADER_FONT = Font(bold=True, color=NAVY_TEXT, size=11)
_THIN = Side(style="thin", color="D6DBE6")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_ALIGN = Alignment(horizontal="left", vertical="top")

# Columns whose prose is long enough to wrap instead of forcing a 60-char-wide column.
WRAP_COLS = {
    "Budget rationale", "Reason", "Test hypothesis", "Notes", "Supporting queries",
    "Primary angles", "Check", "Snippet values", "Workflow note", "Value",
    "Default action", "Description line 1", "Description line 2",
}
# Sensible per-column width caps (chars). Wrapped columns get a fixed comfortable width;
# everything else autosizes within [min, max].
WIDE_WIDTH = 48
NARROW_MAX = 30


# --------------------------------------------------------------------------- helpers
def _ag_name(ag):
    return ag.get("name") or ag.get("ad_group") or ""


def _write_header(ws, headers):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 28


def _style_body(ws, headers):
    """Zebra-band data rows, wrap prose columns, thin grid, freeze header, autofilter, widths.
    Call AFTER all data rows are appended. Row 1 must already be the header row."""
    ncols = len(headers)
    wrap_idx = {i for i, h in enumerate(headers, start=1) if h in WRAP_COLS}
    # Body rows: banding + borders + alignment.
    for r in range(2, ws.max_row + 1):
        band = (r % 2 == 1)  # band odd data rows (row 3, 5, ...) for subtle contrast
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = CELL_BORDER
            cell.alignment = WRAP_ALIGN if c in wrap_idx else TOP_ALIGN
            if band:
                cell.fill = BAND_PATTERN
    # Freeze the header row, add an autofilter over the full used range.
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ncols >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"
    # Column widths.
    for idx in range(1, ncols + 1):
        letter = get_column_letter(idx)
        header = headers[idx - 1] if idx - 1 < len(headers) else ""
        if header in WRAP_COLS:
            ws.column_dimensions[letter].width = WIDE_WIDTH
            continue
        widest = len(str(header))
        for cell in ws[letter]:
            v = cell.value
            if v is None or str(v).startswith("="):
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = max(10, min(NARROW_MAX, widest + 3))


def _display_form(text, match_type):
    mt = (match_type or "").strip().lower()
    if mt == "exact":
        return f"[{text}]"
    if mt == "phrase":
        return f'"{text}"'
    return text


# --------------------------------------------------------------------------- guards
def assert_campaign_consistent(strategy, structuring, rsa_manifest, assets):
    names = {
        "campaign-strategy": strategy.get("campaign"),
        "structuring": structuring.get("campaign"),
        "rsa manifest": rsa_manifest.get("campaign"),
        "assets": assets.get("campaign"),
    }
    present = {k: v for k, v in names.items() if v}
    distinct = set(present.values())
    if len(distinct) > 1:
        raise SystemExit(
            "STOP: campaign name disagrees across input shapes (the human paste between "
            f"phases desynced them): {present}. Reconcile before assembling."
        )


def guard_rsa_paths_and_join(structuring, rsa_manifest):
    """Fail-before-write: every rsa artifact's ads.json must exist, and every RSA ad_group
    must match a structuring ad-group name (else ads.csv references a non-existent ad group)."""
    missing = [a.get("ads_json") for a in rsa_manifest.get("rsa_artifacts", [])
               if not a.get("ads_json") or not os.path.exists(a.get("ads_json"))]
    if missing:
        raise SystemExit(
            "STOP: rsa manifest references ads.json file(s) that do not exist (moved file or a "
            "relative path run from the wrong cwd). Cannot assemble an ad-less campaign:\n  "
            + "\n  ".join(str(m) for m in missing)
        )
    known = {_ag_name(ag) for ag in structuring.get("ad_groups", [])}
    rsa_ags = set()
    orphans = []
    for art in rsa_manifest.get("rsa_artifacts", []):
        with open(art["ads_json"], encoding="utf-8") as f:
            ag = json.load(f).get("ad_group") or art.get("ad_group", "")
        if ag:
            rsa_ags.add(ag)
        if ag and ag not in known:
            orphans.append(ag)
    if orphans:
        # Hard-exit: ads.csv would reference an ad group adgroups.csv never creates.
        raise SystemExit(
            "STOP: RSA ad_group(s) do not match any structuring ad-group name (ads.csv would "
            f"reference an ad group adgroups.csv never creates): {sorted(set(orphans))}. "
            f"Known ad groups: {sorted(known)}."
        )
    # Reverse direction is NOT a hard error (operator may build keywords first, ads later):
    # a structuring ad group with no RSA ships as a keywords-only group. Surface it as a
    # Must-pass launch gate in tab 08 instead of silently shipping a non-functional group.
    return sorted(known - rsa_ags)  # ad-less ad groups


def guard_positive_match_types(structuring):
    """Guard 1: every positive keyword has an explicit Exact or Phrase. No blank/Broad."""
    bad = []
    for ag in structuring.get("ad_groups", []):
        for kw in ag.get("keywords", []):
            mt = (kw.get("match_type") or "").strip().lower()
            if mt not in ("exact", "phrase"):
                bad.append(f"{_ag_name(ag)} / '{kw.get('text')}' -> '{kw.get('match_type')}'")
    if bad:
        raise SystemExit(
            "STOP: positive keyword rows with blank/Broad match type (would create Broad "
            "keywords on import, violating the no-Broad lock). Fix in structuring:\n  "
            + "\n  ".join(bad)
        )


# --------------------------------------------------------------------------- workbook tabs
def tab_readme(wb, strategy, meta):
    ws = wb.active
    ws.title = "00 README"
    headers = ["Field", "Value"]
    _write_header(ws, headers)
    rows = [
        ("Deliverable", meta.get("deliverable", "Google Ads campaign skeleton (campaign-build)")),
        ("Date", meta.get("date", "")),
        ("Account", strategy.get("account_id", "")),
        ("Campaign", strategy.get("campaign", "")),
        ("Important launch gate", strategy.get("tracking_prerequisite", "")),
        ("Workflow note", "This workbook is for review and client confirmation. After it is "
                          "approved, it is converted to Google Ads Editor CSVs (separate step) "
                          "and imported manually. Nothing is pushed to the account automatically."),
    ]
    for r in rows:
        ws.append(list(r))
    # README has a wide free-text "Value" column — give it room and wrap.
    _style_body(ws, headers)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80


def tab_campaign_settings(wb, strategy):
    """Mirror Ian's tab 01: horizontal columns, header row + one data row. Budget is split
    into a numeric 'Daily budget (DKK)' cell + a 'Budget rationale' cell so the number the
    CSV converter needs is never lost behind the prose rationale."""
    ws = wb.create_sheet("01 Campaign settings")
    nets = strategy.get("networks") or {}
    budget = strategy.get("budget_recommendation") or {}
    net_str = "; ".join(
        n for n, on in [("Search", nets.get("search")),
                        ("Search Partners", nets.get("search_partners")),
                        ("Display", nets.get("display"))] if on
    )
    # (header, value) in Ian's exact column order.
    pairs = [
        ("Account ID", strategy.get("account_id", "")),
        ("Campaign", strategy.get("campaign", "")),
        ("Campaign type", strategy.get("campaign_type", "Search")),
        ("Campaign state", strategy.get("campaign_state", "Paused")),
        ("Goal", strategy.get("goal", "Leads")),
        # Two cells, not one: the numeric daily budget is what campaigns.csv needs, so it must
        # survive on its own and never be overwritten by the prose rationale (gap fix).
        ("Daily budget (DKK)", budget.get("daily_dkk", "")),
        ("Budget rationale", budget.get("rationale", "")),
        ("Bidding strategy", strategy.get("bidding_strategy", "")),
        ("Primary conversion action", strategy.get("primary_conversion_action", "")),
        ("Do not optimize toward", strategy.get("do_not_optimize_toward", "")),
        ("Location", strategy.get("location", "")),
        ("Location option", strategy.get("location_option", "")),
        ("Languages", "; ".join(strategy.get("languages") or [])),
        ("Networks", net_str),
        ("Ad rotation", strategy.get("ad_rotation", "")),
        ("Start match types", strategy.get("start_match_types", "")),
        ("AI Max for Search", strategy.get("ai_max_for_search", "")),
        ("Target CPA switch rule", strategy.get("target_cpa_switch_rule", "")),
        ("Ad schedule", strategy.get("ad_schedule", "")),
        ("Tracking prerequisite", strategy.get("tracking_prerequisite", "")),
    ]
    headers = [h for h, _ in pairs]
    _write_header(ws, headers)
    ws.append([str(v) for _, v in pairs])
    # One wide data row: wrap every cell so long values (rationale, prerequisite) stay readable
    # without a 90-char column. _style_body handles banding/borders/freeze/filter/widths.
    _style_body(ws, headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).alignment = WRAP_ALIGN
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.row_dimensions[2].height = 60


def tab_ad_groups(wb, structuring):
    ws = wb.create_sheet("02 Ad groups")
    # Max CPC is a load-bearing Editor field (adgroups.csv needs it) — it MUST live in the
    # workbook so the Excel→CSV converter is lossless. Blank = let the bid strategy decide.
    cols = ["Campaign", "Ad group", "Intent", "Main kw", "Supporting queries",
            "Final URL", "Path 1", "Path 2", "Max CPC", "Primary angles"]
    _write_header(ws, cols)
    campaign = structuring.get("campaign", "")
    for ag in structuring.get("ad_groups", []):
        kws = ag.get("keywords", [])
        main_kw = kws[0].get("text", "") if kws else ""
        supporting = "; ".join(k.get("text", "") for k in kws[1:])
        paths = ag.get("paths", ["", ""])
        ws.append([
            campaign, _ag_name(ag), ag.get("theme", ""), main_kw, supporting,
            ag.get("landing_page_url", ""),
            paths[0] if len(paths) > 0 else "",
            paths[1] if len(paths) > 1 else "",
            ag.get("max_cpc", ""),
            "; ".join(ag.get("angles", [])),
        ])
    _style_body(ws, cols)


def tab_keywords(wb, structuring):
    ws = wb.create_sheet("03 Keywords")
    cols = ["Campaign", "Ad group", "Keyword", "Match type", "Keyword display",
            "Final URL", "Status", "Notes"]
    _write_header(ws, cols)
    campaign = structuring.get("campaign", "")
    for ag in structuring.get("ad_groups", []):
        for kw in ag.get("keywords", []):
            mt = kw.get("match_type", "")
            ws.append([
                campaign, _ag_name(ag), kw.get("text", ""), mt,
                _display_form(kw.get("text", ""), mt),
                ag.get("landing_page_url", ""),
                "Enabled after launch QA",
                structuring.get("keyword_volume_disclaimer", ""),
            ])
    _style_body(ws, cols)


def tab_negatives(wb, structuring):
    ws = wb.create_sheet("04 Negative keywords")
    # Level + Ad group must be in the workbook: the CSV converter derives Editor's Type
    # (Campaign negative vs ad-group Negative) and the Ad group cell from them. Without these
    # columns a campaign- vs ad-group-level negative is indistinguishable in the Excel.
    cols = ["Campaign", "Level", "Ad group", "Negative keyword", "Match type",
            "Category", "Reason"]
    _write_header(ws, cols)
    campaign = structuring.get("campaign", "")
    negs = structuring.get("negatives") or {}
    shared = negs.get("inherited_shared_list") or {}
    # Reference line ONLY — the 277 terms are applied by reference, never enumerated here.
    if shared.get("apply_by_reference"):
        ws.append([
            campaign, "(shared list)", "",
            f"[SHARED LIST APPLIED BY REFERENCE: '{shared.get('name', SHARED_NEG_NAME)}' "
            f"id {shared.get('shared_set_id', SHARED_NEG_ID)}]",
            "(shared list)", "Inherited MCC list",
            "Attached as a launch-gate step (tab 08), not enumerated. See MCC "
            f"{shared.get('mcc_customer_id', SHARED_NEG_MCC)}.",
        ])
    for n in negs.get("client_specific_additions", []):
        level = n.get("level", "campaign")
        ws.append([
            campaign, level, "" if level == "campaign" else n.get("ad_group", ""),
            n.get("text", ""), n.get("match_type", "broad"),
            n.get("category", ""), n.get("why", ""),
        ])
    _style_body(ws, cols)


def tab_monitor(wb, structuring):
    ws = wb.create_sheet("05 Monitor negatives")
    cols = ["Candidate negative", "Default action", "Reason"]
    _write_header(ws, cols)
    for m in (structuring.get("negatives") or {}).get("monitor_first_candidates") or []:
        ws.append([m.get("text", ""), "Monitor first", m.get("why", "")])
    _style_body(ws, cols)


def _rsa_rows(rsa_manifest):
    """Yield (ad_group, ad_label, final_url, paths, hypothesis, headlines, descriptions)
    for every RSA across the manifest, reading each artifact's ads.json.

    Fail loud on a missing/unresolvable ads.json: a barrier that silently drops an RSA
    and still reports success is exactly the corruption Phase 4 exists to prevent. Resolve
    relative paths against the manifest's own dir is the caller's job; here we hard-error."""
    missing = [a.get("ads_json") for a in rsa_manifest.get("rsa_artifacts", [])
               if not a.get("ads_json") or not os.path.exists(a.get("ads_json"))]
    if missing:
        raise SystemExit(
            "STOP: rsa manifest references ads.json file(s) that do not exist (a moved file "
            "or a relative path run from the wrong cwd). Cannot assemble an ad-less campaign:\n  "
            + "\n  ".join(str(m) for m in missing)
        )
    for art in rsa_manifest.get("rsa_artifacts", []):
        ads_path = art.get("ads_json")
        with open(ads_path, encoding="utf-8") as f:
            data = json.load(f)
        ad_group = data.get("ad_group") or art.get("ad_group", "")
        final_url = data.get("final_url") or art.get("final_url", "")
        ads = data.get("ads") or [data]  # single-RSA form is the object itself
        for i, ad in enumerate(ads, start=1):
            # `or default` (not get(k, default)): a JSON null must also fall back, else
            # len(None) crashes the build after the guards have already passed.
            paths = ad.get("paths") or ["", ""]
            yield (
                ad_group,
                f"RSA {i}" + (f" - {ad.get('vinkel')}" if ad.get("vinkel") else ""),
                ad.get("final_url") or final_url,
                paths,
                ad.get("hypotese") or "",
                ad.get("headlines") or [],
                ad.get("descriptions") or [],
            )


def tab_rsas(wb, structuring, rsa_manifest):
    ws = wb.create_sheet("06 RSAs")
    cols = (["Campaign", "Ad group", "Ad label", "Ad type", "Final URL", "Path 1", "Path 2",
             "Test hypothesis"]
            + [f"Headline {i}" for i in range(1, 16)]
            + [f"Description {i}" for i in range(1, 5)])
    _write_header(ws, cols)
    campaign = structuring.get("campaign", "")
    for ag, label, url, paths, hyp, hls, descs in _rsa_rows(rsa_manifest):
        row = [campaign, ag, label, "Responsive search ad", url,
               paths[0] if len(paths) > 0 else "", paths[1] if len(paths) > 1 else "", hyp]
        row += [(hls[i] if i < len(hls) else "") for i in range(15)]
        row += [(descs[i] if i < len(descs) else "") for i in range(4)]
        ws.append(row)
    _style_body(ws, cols)
    # Headline/description columns hold copy that should wrap and be readable for the client.
    for idx, h in enumerate(cols, start=1):
        if h.startswith(("Headline", "Description")):
            ws.column_dimensions[get_column_letter(idx)].width = 26
            for cell in ws[get_column_letter(idx)]:
                cell.alignment = WRAP_ALIGN


def tab_assets(wb, assets):
    ws = wb.create_sheet("07 Assets")
    # One dedicated column per asset field instead of overloading "Asset text"/"Final URL".
    # The old layout stuffed snippet values into the Final URL cell — fragile for the CSV
    # converter and looks broken on a client-facing sheet. Each row fills only its type's cols;
    # the converter keys off "Asset type". Snippet header + values get their own columns.
    # Level (campaign vs account) must be in the workbook: the converter emits the literal
    # "<Account-level>" in Editor's Campaign column for account-level asset sets (answer 56368).
    # Without this column, account- vs campaign-level attachment is unrecoverable from the Excel.
    cols = ["Campaign", "Level", "Asset type", "Sitelink text", "Final URL",
            "Description line 1", "Description line 2", "Callout text",
            "Snippet header", "Snippet values"]
    _write_header(ws, cols)
    campaign = assets.get("campaign", "")
    level = assets.get("attachment_level", "campaign")
    for s in assets.get("sitelinks", []):
        if s.get("url_source") == "omitted-unconfirmed" or not s.get("final_url"):
            continue  # Firewall B: never ship a guessed/unconfirmed sitelink URL
        ws.append([campaign, level, "Sitelink", s.get("text", ""), s.get("final_url", ""),
                   s.get("desc_line_1", ""), s.get("desc_line_2", ""), "", "", ""])
    for c in assets.get("callouts", []):
        ws.append([campaign, level, "Callout", "", "", "", "", c.get("text", ""), "", ""])
    for sn in assets.get("structured_snippets", []):
        ws.append([campaign, level, "Structured snippet", "", "", "", "", "",
                   sn.get("header", ""), "; ".join(sn.get("values", []))])
    _style_body(ws, cols)


def tab_launch_qa(wb, strategy, structuring, adless_ad_groups=None):
    ws = wb.create_sheet("08 Launch QA")
    qa_cols = ["Priority", "Check", "Owner", "Launch gate"]
    _write_header(ws, qa_cols)
    nets = strategy.get("networks") or {}
    shared = (structuring.get("negatives") or {}).get("inherited_shared_list") or {}
    rows = [
        ("Critical", strategy.get("tracking_prerequisite",
         "HubSpot form conversion fires into Google Ads"), "Tracking/analytics", "Must pass"),
        ("Critical", "Do not optimize toward " + (strategy.get("do_not_optimize_toward")
         or "the unverified conversion") + " until it registers correctly",
         "Tracking/analytics", "Must pass"),
        ("High", "Keyword Planner volume and CPC confirm the daily budget (keywords are "
         "theme-derived, not volume-ranked)", "Paid search", "Should pass"),
        ("High", f"Location targeting is Presence only ({strategy.get('location', '')})",
         "Paid search", "Must pass"),
        ("High", "Search Partners and Display Expansion are off", "Paid search", "Must pass"),
        ("High", f"Attach shared negative list '{shared.get('name', SHARED_NEG_NAME)}' "
         f"(id {shared.get('shared_set_id', SHARED_NEG_ID)}) to the campaign",
         "Paid search", "Must pass"),
        ("High", "All client-specific negatives applied before enabling ads",
         "Paid search", "Must pass"),
        ("Medium", "Final URLs and UTM convention approved (assembler does not emit UTMs yet)",
         "Paid search", "Should pass"),
        ("Medium", "Sitelinks reviewed; confirmed real URLs (none guessed)",
         "Paid search", "Should pass"),
        ("Medium", "Structured-snippet header column verified in Editor (round-trip)",
         "Paid search", "Should pass"),
        ("Post-launch", "Search terms reviewed after 7, 14 and 30 days",
         "Paid search", "Post-launch"),
    ]
    # An ad group with keywords but no RSA is non-functional (no ad serves). Not a hard
    # error (operator may add ads later), but it must NOT ship silently — surface it.
    if adless_ad_groups:
        rows.insert(0, (
            "Critical",
            "Ad group(s) have keywords but NO RSA — no ad will serve. Add an RSA or remove "
            f"the ad group before enabling: {', '.join(adless_ad_groups)}",
            "Paid search", "Must pass",
        ))
    for r in rows:
        ws.append(list(r))
    _style_body(ws, qa_cols)


def tab_validation(wb, rsa_manifest):
    """Guard 2: recompute LEN + Pass independently against the imported limits."""
    ws = wb.create_sheet("09 Validation")
    val_cols = ["Area", "Ad group", "Ad label", "Field", "Text", "Length", "Limit", "Pass"]
    _write_header(ws, val_cols)
    failures = 0
    fail_rows = []  # row numbers to flag red AFTER styling (banding would otherwise overwrite)
    pass_rows = []
    for ag, label, _url, paths, _hyp, hls, descs in _rsa_rows(rsa_manifest):
        items = (
            [("RSA headline", f"Headline {i+1}", t, HEADLINE_LIMIT) for i, t in enumerate(hls)]
            + [("RSA description", f"Description {i+1}", t, DESCRIPTION_LIMIT)
               for i, t in enumerate(descs)]
            + [("Display path", f"Path {i+1}", t, PATH_LIMIT)
               for i, t in enumerate(paths)]
        )
        for area, field, text, limit in items:
            length = len(text or "")
            ok = length <= limit
            if not ok:
                failures += 1
            cell_row = ws.max_row + 1
            ws.append([area, ag, label, field, text, length, limit, "Pass" if ok else "FAIL"])
            (pass_rows if ok else fail_rows).append(cell_row)
    _style_body(ws, val_cols)
    # Re-apply the over-length flags ON TOP of the zebra banding (styling runs first).
    for cell_row in fail_rows:
        ws.cell(row=cell_row, column=8).fill = RED_FILL
        ws.cell(row=cell_row, column=8).font = Font(bold=True, color=RED_TEXT)
        ws.cell(row=cell_row, column=6).fill = RED_FILL
        ws.cell(row=cell_row, column=5).fill = RED_FILL
    for cell_row in pass_rows:
        ws.cell(row=cell_row, column=8).font = Font(color=GREEN_TEXT)
    ws.column_dimensions["E"].width = 50
    for cell in ws["E"]:
        cell.alignment = WRAP_ALIGN
    return failures


# --------------------------------------------------------------------------- overview
def write_overview(path, strategy, structuring, rsa_manifest, assets, date, adless, failures):
    """Emit the 'Kampagne overblik' lead doc: structural facts pre-filled, semantic
    one-liners left as {{model: ...}} slots for the skill to complete (findings are
    semantic, not mechanical). Template + rules: references/kampagne-overblik-template.md.
    ONE lead doc — do not also emit a separate import README."""
    negs = structuring.get("negatives") or {}
    shared = negs.get("inherited_shared_list") or {}
    ags = structuring.get("ad_groups", [])
    n_kw = sum(len(a.get("keywords", [])) for a in ags)
    n_rsa = sum(1 for _ in _rsa_rows(rsa_manifest))
    budget = strategy.get("budget_recommendation") or {}
    lines = [
        f"# Kampagne-overblik — {strategy.get('client', structuring.get('campaign',''))}",
        "",
        f"**Konto:** {strategy.get('account_id','')} · **Kampagne:** "
        f"`{strategy.get('campaign', structuring.get('campaign',''))}` · **Dato:** {date}",
        "**Status ved import:** Paused (aktivér først efter review + launch-QA)",
        "",
        "## Beslutninger (hvad bygget valgte)",
        f"- **Struktur:** {len(ags)} ad groups, {n_kw} keywords (Exact + udvalgt Phrase, ingen "
        f"Broad). {{{{model: én linje struktur-rationale fra structure_rationale}}}}",
        f"- **Negativer:** delt MCC-liste \"{shared.get('name','Generelle negative søgeord')}\" "
        f"påført by-reference (id {shared.get('shared_set_id','6688642473')}) + "
        f"{len(negs.get('client_specific_additions', []))} klient-specifikke. "
        f"{len(negs.get('monitor_first_candidates', []))} monitor-first-kandidater.",
        "- **Keywords:** tema-afledte — validér volumen i Keyword Planner før aktivering.",
        f"- **Annoncer:** {n_rsa} RSA'er.",
        f"- **Assets:** {len(assets.get('sitelinks', []))} sitelinks, "
        f"{len(assets.get('callouts', []))} callouts, "
        f"{len(assets.get('structured_snippets', []))} structured snippets.",
        f"- **Budstrategi/budget:** {strategy.get('bidding_strategy','')} · "
        f"{budget.get('daily_dkk','')} DKK/dag" + (f" — {budget.get('rationale','')}" if budget.get('rationale') else ""),
        "",
        "## Vigtigste fund / flag (læs før go-live)",
    ]
    if adless:
        lines.append(f"- ADVARSEL: Ad group(s) UDEN RSA (ingen annonce serveres): {', '.join(adless)} — "
                     "tilføj RSA eller fjern før aktivering (tab 08).")
    if failures:
        lines.append(f"- ADVARSEL: {failures} felt(er) over hård tegngrænse — se fane 09 Validering (rød). "
                     "Ret før import.")
    lines += [
        "{{model: 2-5 one-liner-fund fra fane 08 + 09 + input-objekterne — fx udeladte "
        "usikre sitelink-URL'er, snippet-header-kolonne UNVERIFIED, tracking-gate, "
        "overlappende ad groups der skal pauses. Hvis intet blokerende: \"Ingen blokerende "
        "fund — klar til review.\"}}",
        "",
        "## Sådan importerer du (bulk-upload til Google Ads Editor)",
        f"1. Åbn **Google Ads Editor**, vælg kontoen ({strategy.get('account_id','')}).",
        "2. **Account → Import → From file** med Editor-CSV'erne. (Editor importerer CSV, ikke "
        ".xlsx — denne workbook er review-laget; Editor-CSV'erne genereres fra den godkendte "
        "workbook af `google-ads-general`-konverteren.)",
        "3. Importér i rækkefølge, kør **Check Changes** efter hver: campaigns → ad groups → "
        "keywords → ads (RSA) → assets → negative keywords.",
        f"4. **Tilknyt den delte negativliste** \"{shared.get('name','Generelle negative søgeord')}\" "
        f"(id {shared.get('shared_set_id','6688642473')}) til kampagnen — den er IKKE i nogen CSV.",
        "5. **Manuelt efter import:** sprog = Dansk, Denmark = Presence (ikke Presence-or-Interest), "
        "verificér leadgen-konverteringshandlingen.",
        "6. Kør **Check Changes**, løs alle røde fejl, verificér status = **Paused**, derefter Post Changes.",
        "",
        "## Før go-live (launch-gate — fuld liste i workbook fane 08)",
        "{{model: Must-pass-rækkerne fra fane 08 som one-liners}}",
        "",
        "---",
        "Genereret af campaign-build assembler. Workbook = detaljen; dette = forsiden. Intet pushet til kontoen.",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# --------------------------------------------------------------------------- main
def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def main():
    ap = argparse.ArgumentParser(description="Phase-4 campaign-build assembler (Excel-only, no API push).")
    ap.add_argument("--strategy", required=True, help="campaign-strategy.json")
    ap.add_argument("--structuring", required=True, help="structuring.json")
    ap.add_argument("--rsa", required=True, help="rsa manifest json")
    ap.add_argument("--assets", required=True, help="assets.json")
    ap.add_argument("--workbook", required=True, help="output .xlsx path")
    ap.add_argument("--overview", default="", help="output 'Kampagne overblik.md' path (optional)")
    ap.add_argument("--date", default="", help="deliverable date (passed in; script can't call now)")
    args = ap.parse_args()

    strategy = load_json(args.strategy)
    structuring = load_json(args.structuring)
    rsa_manifest = load_json(args.rsa)
    assets = load_json(args.assets)

    # Reconcile + guards FIRST (fail loud before writing anything).
    assert_campaign_consistent(strategy, structuring, rsa_manifest, assets)
    guard_positive_match_types(structuring)
    adless_ad_groups = guard_rsa_paths_and_join(structuring, rsa_manifest)

    # Build the 10-tab workbook.
    wb = openpyxl.Workbook()
    tab_readme(wb, strategy, {"date": args.date})
    tab_campaign_settings(wb, strategy)
    tab_ad_groups(wb, structuring)
    tab_keywords(wb, structuring)
    tab_negatives(wb, structuring)
    tab_monitor(wb, structuring)
    tab_rsas(wb, structuring, rsa_manifest)
    tab_assets(wb, assets)
    tab_launch_qa(wb, strategy, structuring, adless_ad_groups)
    failures = tab_validation(wb, rsa_manifest)
    wb.save(args.workbook)

    overview_path = args.overview
    if overview_path:
        write_overview(overview_path, strategy, structuring, rsa_manifest, assets,
                       args.date, adless_ad_groups, failures)

    print(json.dumps({
        "workbook": args.workbook,
        "overview": overview_path or None,
        "validation_failures": failures,
        "adless_ad_groups": adless_ad_groups,
        "note": "Excel-only. The workbook is the client-confirmation artifact; Editor CSVs are "
                "generated later from the confirmed Excel by the google-ads-general converter "
                "skill. No API push. Shared negative list 6688642473 applied by reference "
                "(tab 08 + tab 04 reference line), never enumerated.",
    }, ensure_ascii=False, indent=2))

    if failures:
        print(f"\nWARNING: {failures} field(s) over the hard limit — see tab 09 Validation "
              "(red). Fix before import.", file=sys.stderr)
        sys.exit(3)


if __name__ == "__main__":
    main()
