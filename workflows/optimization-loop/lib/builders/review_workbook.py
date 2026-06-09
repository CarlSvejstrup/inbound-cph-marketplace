#!/usr/bin/env python3
"""Build the optimization-loop REVIEW WORKBOOK (.xlsx) — the editable deliverable.

ARCHITECTURE (locked 2026-06-05): the loop does NOT emit Editor CSVs. It returns ONE
Excel workbook that the ads expert edits, can send to the client, and then hands to a
separate converter skill (in google-ads-general) that does workbook -> Editor CSV. This
mirrors the assembler, which was already made Excel-only (commit eb4ebd9): "the workbook
is a lossless superset; the converter produces the CSVs." Editor imports CSV, not .xlsx
(Google Ads Editor answer 30564) - which is exactly why the converter exists.

THE COLUMN CONTRACT (this is the interface the converter is built against):

Each entity tab splits its columns into two bands:
  1. EDITOR-BOUND columns - exact Google Ads Editor header spelling (answer 57747:
     headers are English, case/space-insensitive). The converter KEEPS these.
  2. METADATA columns - review context (reason, wasted spend, conversions, CPA, source).
     The converter DROPS these. They never go to Editor.

A sentinel row/format is not needed: the converter knows the fixed Editor header set per
entity (documented in SPEC section 3.5) and drops everything else.

### #Original — editing EXISTING entities (the loop's distinguishing need)

The assembler builds NEW campaigns (everything net-new, lands Paused, no #Original). The
loop optimises a LIVE account, so some rows EDIT an existing entity (e.g. an RSA rewrite).
Google Ads Editor uses the `<Column>#Original` convention (answer 57747) to match an edit
to the existing entity and preserve its history instead of creating a duplicate. So:
  - NET-NEW rows (new negative, promoted keyword, brand-new challenger RSA in an ad group
    that had none): no #Original columns. Status as appropriate.
  - EDIT rows (rewriting an existing RSA, changing an existing keyword): include the
    `#Original` column(s) carrying the current live value, so Editor edits in place.
The loop's execute stage decides per row which case applies; this builder writes whatever
#Original cells it is given and the converter preserves any `*#Original` column verbatim.

Self-bootstraps openpyxl. Runs locally and in an agent (the workflow JS itself cannot run
Python - see SPEC section 2).
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)


_ensure_openpyxl()
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# --- styling (matches the review-sheet house look: blue header, wrap) ---
HEADER_BG = "1F4E78"
META_BG = "D9E1F2"          # metadata columns get a lighter header so the expert sees the divide
HEADER_FONT = Font(bold=True, color="FFFFFF")
META_FONT = Font(bold=True, color="1F4E78")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
_MATCH = {"EXACT": "Exact", "PHRASE": "Phrase", "BROAD": "Broad"}


def _match(raw):
    return _MATCH.get((raw or "").strip().upper(), raw or "")


def _fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")


def _sheet(wb, title, editor_headers, meta_headers, rows, widths=None):
    """Write one entity tab. editor_headers get the dark header (converter KEEPS);
    meta_headers get the light header (converter DROPS). rows are dicts keyed by header."""
    ws = wb.create_sheet(title)
    headers = editor_headers + meta_headers
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.alignment = WRAP
        if h in editor_headers:
            cell.fill = _fill(HEADER_BG)
            cell.font = HEADER_FONT
        else:
            cell.fill = _fill(META_BG)
            cell.font = META_FONT
    for r, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=row.get(h, "")).alignment = WRAP
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


# Editor-bound header sets per entity (the converter's KEEP list). Exact Editor spelling.
KEYWORDS_EDITOR = ["Campaign", "Ad group", "Keyword", "Match type", "Status"]
# Negatives speak the SAME vocabulary as the assembler's tab 04 so the converter's
# build_negatives() works unchanged for both workbooks (one contract, no fork):
# Campaign + Level + Ad group + Negative keyword + Match type. The converter derives
# Editor's Type (Campaign negative vs Negative) from Level, never from Match type.
NEGATIVES_EDITOR = ["Campaign", "Level", "Ad group", "Negative keyword", "Match type"]
RSA_EDITOR = (["Campaign", "Ad group", "Ad type", "Final URL", "Path 1", "Path 2"]
              + [f"Headline {i}" for i in range(1, 16)]
              + [f"Description {i}" for i in range(1, 5)]
              + ["Status"])


def _readme_sheet(wb, client, account_id, period, today, account_level_notes=None):
    ws = wb.create_sheet("Laes mig", 0)
    ws["A1"] = f"Optimerings-forslag — {client}"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        f"Konto: {account_id}    Periode: {period}    Genereret: {today}",
        "",
        "SAADAN BRUGER DU FILEN:",
        "1. Gennemgaa hver fane. De MOERKEBLAA kolonner er Google Ads Editor-felter.",
        "   De LYSEBLAA kolonner er kontekst til dig (begrundelse, spild, konverteringer) —",
        "   de bliver IKKE importeret.",
        "2. Ret frit: slet raekker du er uenig i, juster tekst, tilpas bud.",
        "3. Alle RSA-forslag er NYE challengers (status Paused) — ikke rettelser af eksisterende",
        "   annoncer. Saet den gamle annonce paa pause/fjern den FOERST naar challengeren har",
        "   vist sig bedre. (At redigere en live RSA nulstiller dens laering.)",
        "4. Naar du er faerdig: koer konverterings-skillet (workbook -> Editor-CSV), og",
        "   importer CSV'erne i Google Ads Editor. Gennemgaa groen/gul diff, tryk Send.",
        "",
        "Intet i denne fil er skrevet til kontoen. Du har fuld kontrol.",
    ]
    # Account-level negatives can't be a CSV row (Editor only has campaign/ad-group level).
    # They are fanned out to one Campaign-negative row per active campaign on the Negative
    # keywords tab. Surface the cleaner alternative (a shared negative list) here so the
    # expert can choose it instead of the fanned rows.
    if account_level_notes:
        lines += [
            "",
            "KONTO-NIVEAU NEGATIVE (vigtigt):",
            "Editor kan ikke importere konto-niveau negative via CSV. Foelgende ord er derfor",
            "udfoldet til en 'Campaign negative'-raekke PER aktiv kampagne paa fanen 'Negative",
            "keywords'. Alternativt (renere): tilfoej dem i stedet til en delt negativliste i",
            "Google Ads UI'en og tilknyt den til kontoens kampagner. Vaelg det ene ELLER det andet:",
        ]
        for n in account_level_notes:
            lines.append(f"   - {n.get('keyword', '')}  ({n.get('reason', '')})")
    for i, ln in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=ln)
    ws.column_dimensions["A"].width = 90
    return ws


def build(data, out_path):
    """Build the review workbook.

    data schema:
    {
      "client", "account_id", "period", "today",
      "active_campaigns": ["string"],   # active campaign names; account-level negatives fan out across these
      "negatives": [ {keyword, match_type (EXACT|PHRASE), level (ad_group|campaign|account),
                      campaign, ad_group, wasted_spend_dkk, reason} ],
      "winners":   [ {term, campaign, ad_group, conversions, cpa_dkk, reason} ],   # promote->Exact, Paused
      "rsa_rows":  [ {campaign, ad_group, headlines[], descriptions[], paths[2], final_url,
                      status (Paused), reason} ]
    }
    Every RSA row is a NET-NEW challenger (Paused). The loop never edits a live RSA in place
    (resets learning + Editor CSV can't reliably match an RSA) — see the RSA tab comment.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    client = data.get("client", "")
    account_id = data.get("account_id", "")
    period = data.get("period", "")
    today = data.get("today", "")

    # --- Negative keywords tab ---
    # Google Ads Editor's CSV import supports ONLY campaign- and ad-group-level negatives
    # (Type = "Campaign negative" / "Negative", answer 57747). There is NO account-level
    # negative row in the CSV namespace; account-wide blocking is a shared-negative-list
    # (shared set) operation done in the UI. So an ACCOUNT-level finding is fanned out here
    # into one "Campaign negative" row per active campaign (same account-wide blocking effect,
    # fully importable) AND listed once on the Laes mig tab with the shared-list instruction.
    # The fan-out lives in the builder (not the converter) because the builder has the active
    # campaign list and because the workbook should be what-you-see-is-what-imports: the expert
    # reviews the actual per-campaign rows that will be created, not a single row that secretly
    # explodes at CSV time. Decision 2026-06-09.
    active_campaigns = data.get("active_campaigns") or []
    account_level_notes = []
    neg_rows = []
    for n in data.get("negatives", []):
        level = (n.get("level") or "campaign").lower()
        if level == "account":
            account_level_notes.append(n)
            targets = active_campaigns or [n.get("campaign", "")]
            for camp in targets:
                neg_rows.append({
                    "Campaign": camp,
                    "Level": "campaign",   # fanned to campaign-level rows
                    "Ad group": "",
                    "Negative keyword": n.get("keyword", ""),
                    "Match type": _match(n.get("match_type", "")),
                    "Niveau (oprindeligt)": "account -> fanned to campaign",
                    "Spildt budget (DKK)": n.get("wasted_spend_dkk", ""),
                    "Begrundelse": n.get("reason", ""),
                })
            continue
        neg_rows.append({
            "Campaign": n.get("campaign", ""),
            "Level": level,
            "Ad group": n.get("ad_group", "") if level == "ad_group" else "",
            "Negative keyword": n.get("keyword", ""),
            "Match type": _match(n.get("match_type", "")),
            # metadata (dropped by converter):
            "Niveau (oprindeligt)": level,
            "Spildt budget (DKK)": n.get("wasted_spend_dkk", ""),
            "Begrundelse": n.get("reason", ""),
        })
    # Readme first (inserted at index 0). Built here so it can list the account-level notes
    # computed in the negatives loop above.
    _readme_sheet(wb, client, account_id, period, today, account_level_notes)

    _sheet(wb, "Negative keywords", NEGATIVES_EDITOR,
           ["Niveau (oprindeligt)", "Spildt budget (DKK)", "Begrundelse"], neg_rows,
           widths=[26, 10, 22, 28, 12, 22, 18, 60])

    # --- Keyword expansion tab (promote winners to Exact, Paused) ---
    kw_rows = []
    for w in data.get("winners", []):
        kw_rows.append({
            "Campaign": w.get("campaign", ""),
            "Ad group": w.get("ad_group", ""),
            "Keyword": w.get("term", ""),
            "Match type": "Exact",
            "Status": "Paused",
            # metadata:
            "Konverteringer": w.get("conversions", ""),
            "CPA (DKK)": w.get("cpa_dkk", ""),
            "Begrundelse": w.get("reason", ""),
        })
    _sheet(wb, "Nye keywords (vindere)", KEYWORDS_EDITOR,
           ["Konverteringer", "CPA (DKK)", "Begrundelse"], kw_rows,
           widths=[26, 22, 28, 12, 10, 13, 11, 60])

    # --- RSA challengers tab ---
    # EVERY RSA change is a NET-NEW challenger (Paused), never an in-place edit. Rationale
    # (decision 2026-06-09): editing a live RSA's creative resets its learning — RSAs are
    # effectively immutable in Google Ads, so "improving" one is a new-ad operation under the
    # hood (SPEC §6.4 + the ASA/Inbound memory note). Google Ads Editor's docs do not even
    # confirm RSA text-edit-in-place via CSV, and never enumerate which #Original fields would
    # match an RSA — so an #Original RSA edit row risks BOTH a silent duplicate (no match) and
    # clobbered headlines 2-15 (treated as full new content). The safe, idiomatic move is a
    # fresh challenger: the human pauses/removes the old ad after the challenger proves out.
    # (The converter still supports #Original passthrough for genuinely-editable entities like
    # keyword bid/URL — it is just never emitted for RSAs.)
    rsa_editor_headers = list(RSA_EDITOR)
    rsa_rows = []
    for r in data.get("rsa_rows", []):
        row = {
            "Campaign": r.get("campaign", ""),
            "Ad group": r.get("ad_group", ""),
            "Ad type": "Responsive search ad",
            "Final URL": r.get("final_url", ""),
            "Path 1": (r.get("paths") or ["", ""])[0],
            "Path 2": (r.get("paths") or ["", ""])[1] if len(r.get("paths") or []) > 1 else "",
            "Status": r.get("status", "Paused"),
        }
        for i, h in enumerate(r.get("headlines", [])[:15], start=1):
            row[f"Headline {i}"] = h
        for i, d in enumerate(r.get("descriptions", [])[:4], start=1):
            row[f"Description {i}"] = d
        row["Begrundelse"] = r.get("reason", "")
        rsa_rows.append(row)
    _sheet(wb, "RSA challengers", rsa_editor_headers, ["Begrundelse"], rsa_rows)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import argparse
    import json
    ap = argparse.ArgumentParser(description="Build the optimization-loop review workbook (.xlsx).")
    ap.add_argument("--in", dest="inp", required=True, help="path to the findings JSON")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()
    build(json.loads(Path(args.inp).read_text()), args.out)
    print(f"Wrote {args.out}")
